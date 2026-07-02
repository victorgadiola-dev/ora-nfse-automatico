from __future__ import annotations

from datetime import date, datetime, timezone
from email.utils import parsedate_to_datetime
import os
import re
import time
from decimal import Decimal
from html import escape, unescape
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urlencode

from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.adn_client import AdnClient, AdnClientError, extract_xml_payloads
from app.cert_utils import CertificadoErro, extract_cnpj_root_from_pfx, sanitize_filename
from app.conferencia_excel import EXPECTED_COLUMNS, compare_excel_with_notes, parse_money as parse_excel_money, read_excel_rows
from app.config import get_settings
from app.nfse_parser import ParsedNfse, define_papel_cliente, digits, expand_upload, parse_nfses_from_xml_bytes
from app.secret_store import SecretStoreError, decrypt_secret, encrypt_secret, get_or_create_secret_key
from app.store import JsonStore, store, utc_now_iso

app = FastAPI(
    title="ORA NFS-e Automático",
    version="0.10.0",
    description="Busca automática de NFS-e no ADN/NFS-e Nacional, sem banco de dados, com painel ORA, relatórios por competência/emissão, conferência de Excel importado e interface em padrão de sistema operacional.",
)

STATIC_DIR = Path(__file__).resolve().parent / "static"
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.on_event("startup")
def startup() -> None:
    get_settings().ensure_dirs()
    get_or_create_secret_key()
    store.read()


# ---------------------------------------------------------------------------
# Utilidades de formatação e cálculo
# ---------------------------------------------------------------------------

MONEY_FIELDS = [
    "valor_servico",
    "valor_deducoes",
    "valor_desconto_condicionado",
    "valor_desconto_incondicionado",
    "base_calculo",
    "valor_iss",
    "valor_iss_retido",
    "valor_pis",
    "valor_cofins",
    "valor_inss",
    "valor_ir",
    "valor_csll",
    "retencao_pis_cofins_csll_base",
    "valor_pis_apurado",
    "valor_cofins_apurado",
    "outras_retencoes",
    "total_descontos",
    "total_retencoes_federais",
    "total_retencoes",
    "valor_liquido",
]

DETAIL_COLUMNS = [
    ("competencia", "Competência"),
    ("data_emissao", "Emissão"),
    ("papel_cliente", "Papel"),
    ("status", "Status"),
    ("numero", "Número"),
    ("serie", "Série"),
    ("chave_acesso", "Chave"),
    ("razao_prestador", "Prestador"),
    ("cnpj_prestador", "CNPJ prestador"),
    ("razao_tomador", "Tomador"),
    ("cnpj_tomador", "CNPJ tomador"),
    ("municipio_prestacao", "Município prestação"),
    ("codigo_servico", "Cód. serviço"),
    ("item_lista_servico", "Item lista"),
    ("codigo_tributacao_municipio", "Tributação município"),
    ("codigo_cnae", "CNAE"),
    ("discriminacao", "Discriminação"),
    ("valor_servico", "Valor serviços"),
    ("valor_deducoes", "Deduções"),
    ("valor_desconto_condicionado", "Desc. cond."),
    ("valor_desconto_incondicionado", "Desc. incond."),
    ("base_calculo", "Base ISS"),
    ("aliquota_iss", "Alíquota ISS"),
    ("valor_iss", "ISS"),
    ("iss_retido", "ISS retido?"),
    ("iss_retido_tipo", "Tipo retenção ISS"),
    ("valor_iss_retido", "ISS retido"),
    ("valor_pis", "PIS retido"),
    ("valor_cofins", "COFINS retido"),
    ("valor_csll", "CSLL retida"),
    ("valor_ir", "IRRF retido"),
    ("valor_inss", "INSS/CP retido"),
    ("retencao_pis_cofins_csll_tipo", "Tipo PIS/COFINS/CSLL"),
    ("retencao_pis_cofins_csll_base", "vRetCSLL agregado"),
    ("retencao_pis_cofins_csll_criterio", "Critério retenção social"),
    ("valor_pis_apurado", "PIS apurado XML"),
    ("valor_cofins_apurado", "COFINS apurado XML"),
    ("outras_retencoes", "Outras ret."),
    ("total_retencoes_federais", "Ret. federais"),
    ("total_retencoes", "Total retido"),
    ("valor_liquido", "Valor líquido"),
    ("origem", "Origem"),
    ("nsu", "NSU"),
]


def now_iso() -> str:
    return utc_now_iso()


def dec(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    return Decimal(str(value)).quantize(Decimal("0.01"))


def fmt_money(value: Any) -> str:
    number = dec(value)
    return f"R$ {number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_num(value: Any) -> str:
    number = dec(value)
    return f"{number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_date(value: str | date | None) -> str:
    if not value:
        return ""
    if isinstance(value, date):
        parsed = value
    else:
        try:
            parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00")).date()
        except ValueError:
            return str(value)
    return parsed.strftime("%d/%m/%Y")


def parse_date_filter(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).date()
    except ValueError:
        return None


def mask_cnpj(value: str | None) -> str:
    d = digits(value) or ""
    if len(d) != 14:
        return value or ""
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"


def status_class(status: str | None) -> str:
    s = (status or "").upper()
    if "CANCEL" in s:
        return "err"
    if "SUBSTIT" in s or "PEND" in s:
        return "warn"
    return "ok"


def pill(text: str | None, kind: str | None = None) -> str:
    value = text or "-"
    return f"<span class='pill {escape(kind or status_class(value))}'>{escape(value)}</span>"


def iss_retido_tipo_label(value: Any) -> str:
    code = str(value or "").strip()
    labels = {
        "1": "1 · não retido",
        "2": "2 · retido pelo tomador",
        "3": "3 · retido pelo intermediário",
    }
    return labels.get(code, code or "-")


def social_retencao_tipo_label(value: Any) -> str:
    code = str(value or "").strip()
    labels = {
        "0": "0 · PIS/COFINS/CSLL não retidos",
        "1": "1 · PIS/COFINS retidos",
        "2": "2 · PIS/COFINS não retidos",
        "3": "3 · PIS/COFINS/CSLL retidos",
        "4": "4 · PIS/COFINS retidos",
        "5": "5 · PIS retido",
        "6": "6 · COFINS retido",
        "7": "7 · COFINS/CSLL retidos",
        "8": "8 · CSLL retida",
        "9": "9 · PIS/CSLL retidos",
    }
    return labels.get(code, code or "-")


DATE_BASE_OPTIONS = [
    ("competencia", "Competência"),
    ("emissao", "Data de emissão"),
]


def normalize_date_base(value: Any) -> str:
    normalized = re.sub(r"[^a-z0-9]", "", str(value or "").lower())
    if normalized in {"emissao", "dataemissao", "dtemissao", "emissaonfse"}:
        return "emissao"
    return "competencia"


def date_base_field(value: Any) -> str:
    return "data_emissao" if normalize_date_base(value) == "emissao" else "competencia"


def date_base_label(value: Any) -> str:
    return "data de emissão" if normalize_date_base(value) == "emissao" else "competência"


def note_date_for_base(note: dict[str, Any], data_base: Any = "competencia") -> date | None:
    field = date_base_field(data_base)
    return parse_date_filter(str(note.get(field) or ""))


def date_base_options_html(selected: Any) -> str:
    return html_select_options(DATE_BASE_OPTIONS, normalize_date_base(selected))


def cliente_label(cliente: dict[str, Any] | None) -> str:
    if not cliente:
        return "Todos os clientes"
    return f"{cliente.get('razao_social','')} · {mask_cnpj(cliente.get('cnpj'))}"


def find_by_id(items: list[dict[str, Any]], item_id: int | str | None) -> dict[str, Any] | None:
    if item_id in (None, ""):
        return None
    try:
        target = int(item_id)
    except (TypeError, ValueError):
        return None
    for item in items:
        if int(item.get("id", -1)) == target:
            return item
    return None


def find_cliente_by_cnpj(data: dict[str, Any], cnpj: str | None) -> dict[str, Any] | None:
    d = digits(cnpj)
    for cliente in data.get("clientes", []):
        if digits(cliente.get("cnpj")) == d:
            return cliente
    return None


def should_count_active(note: dict[str, Any]) -> bool:
    return str(note.get("status") or "").upper() == "AUTORIZADA"


def should_count_faturamento(note: dict[str, Any]) -> bool:
    return should_count_active(note) and note.get("papel_cliente") == "PRESTADOR"


def note_month(note: dict[str, Any], data_base: Any = "competencia") -> str:
    selected_date = note_date_for_base(note, data_base)
    return selected_date.strftime("%Y-%m") if selected_date else "Sem data"


def filtered_notes(
    data: dict[str, Any],
    cliente_id: int | None = None,
    inicio: date | None = None,
    fim: date | None = None,
    papel: str = "PRESTADOR",
    status: str = "TODOS",
    data_base: Any = "competencia",
) -> list[dict[str, Any]]:
    cliente = find_by_id(data.get("clientes", []), cliente_id)
    cnpj_cliente = digits(cliente.get("cnpj")) if cliente else None
    papel = (papel or "PRESTADOR").upper()
    status = (status or "TODOS").upper()
    base = normalize_date_base(data_base)
    field = date_base_field(base)

    result: list[dict[str, Any]] = []
    for note in data.get("notas", []):
        if cnpj_cliente and digits(note.get("cnpj_cliente")) != cnpj_cliente:
            continue
        if papel != "TODOS" and note.get("papel_cliente") != papel:
            continue
        if status != "TODOS" and str(note.get("status") or "").upper() != status:
            continue
        selected_date = note_date_for_base(note, base)
        if (inicio or fim) and selected_date is None:
            continue
        if inicio and selected_date and selected_date < inicio:
            continue
        if fim and selected_date and selected_date > fim:
            continue
        result.append(note)
    result.sort(
        key=lambda n: (
            str(n.get(field) or ""),
            str(n.get("competencia") or ""),
            str(n.get("data_emissao") or ""),
            str(n.get("numero") or ""),
        ),
        reverse=True,
    )
    return result


def totals_for_notes(notes: list[dict[str, Any]]) -> dict[str, Any]:
    active = [n for n in notes if should_count_active(n)]
    faturamento = [n for n in notes if should_count_faturamento(n)]
    tomadas = [n for n in notes if should_count_active(n) and n.get("papel_cliente") == "TOMADOR"]

    def sum_field(group: list[dict[str, Any]], field: str) -> Decimal:
        return sum((dec(n.get(field)) for n in group), Decimal("0.00"))

    return {
        "qtd_total": len(notes),
        "qtd_autorizadas": len(active),
        "qtd_canceladas": sum(1 for n in notes if str(n.get("status") or "").upper() == "CANCELADA"),
        "qtd_substituidas": sum(1 for n in notes if str(n.get("status") or "").upper() == "SUBSTITUIDA"),
        "qtd_emitidas": sum(1 for n in notes if n.get("papel_cliente") == "PRESTADOR"),
        "qtd_tomadas": sum(1 for n in notes if n.get("papel_cliente") == "TOMADOR"),
        "faturamento_bruto": sum_field(faturamento, "valor_servico"),
        "faturamento_liquido": sum_field(faturamento, "valor_liquido"),
        "retencoes_faturamento": sum_field(faturamento, "total_retencoes"),
        "iss_retido_faturamento": sum_field(faturamento, "valor_iss_retido"),
        "servicos_tomados": sum_field(tomadas, "valor_servico"),
        "valor_servico": sum_field(active, "valor_servico"),
        "valor_deducoes": sum_field(active, "valor_deducoes"),
        "total_descontos": sum_field(active, "total_descontos"),
        "base_calculo": sum_field(active, "base_calculo"),
        "valor_iss": sum_field(active, "valor_iss"),
        "valor_iss_retido": sum_field(active, "valor_iss_retido"),
        "valor_pis": sum_field(active, "valor_pis"),
        "valor_cofins": sum_field(active, "valor_cofins"),
        "valor_inss": sum_field(active, "valor_inss"),
        "valor_ir": sum_field(active, "valor_ir"),
        "valor_csll": sum_field(active, "valor_csll"),
        "retencao_pis_cofins_csll_base": sum_field(active, "retencao_pis_cofins_csll_base"),
        "valor_pis_apurado": sum_field(active, "valor_pis_apurado"),
        "valor_cofins_apurado": sum_field(active, "valor_cofins_apurado"),
        "outras_retencoes": sum_field(active, "outras_retencoes"),
        "total_retencoes_federais": sum_field(active, "total_retencoes_federais"),
        "total_retencoes": sum_field(active, "total_retencoes"),
        "valor_liquido": sum_field(active, "valor_liquido"),
    }


RETENTION_TOTAL_FIELDS = [
    "valor_iss_retido",
    "valor_pis",
    "valor_cofins",
    "valor_inss",
    "valor_ir",
    "valor_csll",
    "retencao_pis_cofins_csll_base",
    "outras_retencoes",
    "total_retencoes_federais",
    "total_retencoes",
]


def sum_field_notes(notes: list[dict[str, Any]], field: str) -> Decimal:
    return sum((dec(n.get(field)) for n in notes), Decimal("0.00"))


def totals_for_role(notes: list[dict[str, Any]], papel_cliente: str) -> dict[str, Any]:
    papel_cliente = papel_cliente.upper()
    role_all = [n for n in notes if str(n.get("papel_cliente") or "").upper() == papel_cliente]
    group = [n for n in role_all if should_count_active(n)]
    return {
        "qtd_total": len(role_all),
        "qtd": len(group),
        "qtd_canceladas": sum(1 for n in role_all if str(n.get("status") or "").upper() == "CANCELADA"),
        "qtd_substituidas": sum(1 for n in role_all if str(n.get("status") or "").upper() == "SUBSTITUIDA"),
        "valor_servico": sum_field_notes(group, "valor_servico"),
        "valor_deducoes": sum_field_notes(group, "valor_deducoes"),
        "total_descontos": sum_field_notes(group, "total_descontos"),
        "base_calculo": sum_field_notes(group, "base_calculo"),
        "valor_iss": sum_field_notes(group, "valor_iss"),
        "valor_iss_retido": sum_field_notes(group, "valor_iss_retido"),
        "valor_pis": sum_field_notes(group, "valor_pis"),
        "valor_cofins": sum_field_notes(group, "valor_cofins"),
        "valor_inss": sum_field_notes(group, "valor_inss"),
        "valor_ir": sum_field_notes(group, "valor_ir"),
        "valor_csll": sum_field_notes(group, "valor_csll"),
        "retencao_pis_cofins_csll_base": sum_field_notes(group, "retencao_pis_cofins_csll_base"),
        "valor_pis_apurado": sum_field_notes(group, "valor_pis_apurado"),
        "valor_cofins_apurado": sum_field_notes(group, "valor_cofins_apurado"),
        "outras_retencoes": sum_field_notes(group, "outras_retencoes"),
        "total_retencoes_federais": sum_field_notes(group, "total_retencoes_federais"),
        "total_retencoes": sum_field_notes(group, "total_retencoes"),
        "valor_liquido": sum_field_notes(group, "valor_liquido"),
    }


def resumo_empresas_rows(
    data: dict[str, Any],
    inicio: date | None = None,
    fim: date | None = None,
    status: str = "TODOS",
    cliente_ids: list[int] | None = None,
    data_base: Any = "competencia",
) -> list[dict[str, Any]]:
    selected = set(normalize_id_list(cliente_ids))
    rows: list[dict[str, Any]] = []
    for cliente in sorted(data.get("clientes", []), key=lambda c: c.get("razao_social", "")):
        cid = int(cliente["id"])
        if selected and cid not in selected:
            continue
        notes = filtered_notes(data, cliente_id=cid, inicio=inicio, fim=fim, papel="TODOS", status=status, data_base=data_base)
        prestado = totals_for_role(notes, "PRESTADOR")
        tomado = totals_for_role(notes, "TOMADOR")
        rows.append({
            "cliente_id": cliente.get("id"),
            "razao_social": cliente.get("razao_social", ""),
            "cnpj": cliente.get("cnpj", ""),
            "municipios_notas": municipios_from_notes(notes),
            "qtd_total": len(notes),
            "qtd_canceladas": sum(1 for n in notes if str(n.get("status") or "").upper() == "CANCELADA"),
            "qtd_substituidas": sum(1 for n in notes if str(n.get("status") or "").upper() == "SUBSTITUIDA"),
            "prestado": prestado,
            "tomado": tomado,
        })
    return rows


def resumo_empresas_totais(rows: list[dict[str, Any]]) -> dict[str, Any]:
    total = {
        "qtd_total": sum(int(r.get("qtd_total") or 0) for r in rows),
        "qtd_canceladas": sum(int(r.get("qtd_canceladas") or 0) for r in rows),
        "qtd_substituidas": sum(int(r.get("qtd_substituidas") or 0) for r in rows),
        "prestado": {},
        "tomado": {},
    }
    for papel in ["prestado", "tomado"]:
        for key in ["qtd_total", "qtd", "qtd_canceladas", "qtd_substituidas", "valor_servico", "valor_deducoes", "total_descontos", "base_calculo", "valor_iss", "valor_iss_retido", "valor_pis", "valor_cofins", "valor_inss", "valor_ir", "valor_csll", "retencao_pis_cofins_csll_base", "valor_pis_apurado", "valor_cofins_apurado", "outras_retencoes", "total_retencoes_federais", "total_retencoes", "valor_liquido"]:
            if key in {"qtd_total", "qtd", "qtd_canceladas", "qtd_substituidas"}:
                total[papel][key] = sum(int(r[papel].get(key) or 0) for r in rows)
            else:
                total[papel][key] = sum((dec(r[papel].get(key)) for r in rows), Decimal("0.00"))
    return total


def render_value_cell(value: Any, key: str | None = None) -> str:
    if key in MONEY_FIELDS or isinstance(value, Decimal):
        return f"<td class='num'>{fmt_money(value)}</td>"
    return f"<td class='num'>{escape(str(value or ''))}</td>"


def monthly_rows(notes: list[dict[str, Any]], data_base: Any = "competencia") -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for note in notes:
        grouped.setdefault(note_month(note, data_base), []).append(note)
    rows: list[dict[str, Any]] = []
    for mes, group in grouped.items():
        totals = totals_for_notes(group)
        rows.append({"mes": mes, **totals})
    rows.sort(key=lambda row: row["mes"], reverse=True)
    return rows


def html_select_options(values: list[tuple[str | int, str]], selected: str | int | None) -> str:
    selected_str = "" if selected is None else str(selected)
    return "".join(
        f"<option value='{escape(str(value))}' {'selected' if str(value) == selected_str else ''}>{escape(label)}</option>"
        for value, label in values
    )


def normalize_id_list(values: list[int] | tuple[int, ...] | None) -> list[int]:
    if not values:
        return []
    result: list[int] = []
    for value in values:
        try:
            item = int(value)
        except (TypeError, ValueError):
            continue
        if item not in result:
            result.append(item)
    return result


def build_query(**params: Any) -> str:
    clean: dict[str, Any] = {}
    for key, value in params.items():
        if value in (None, "", []):
            continue
        clean[key] = value
    return urlencode(clean, doseq=True)


def periodo_label(inicio: str | None, fim: str | None, data_base: Any | None = None) -> str:
    if inicio and fim:
        text = f"{fmt_date(inicio)} a {fmt_date(fim)}"
    elif inicio:
        text = f"a partir de {fmt_date(inicio)}"
    elif fim:
        text = f"até {fmt_date(fim)}"
    else:
        text = "todo o histórico disponível"
    if data_base is not None:
        text += f" · por {date_base_label(data_base)}"
    return text


def clean_response_preview(value: str | None, limit: int = 180) -> str:
    """Remove HTML e ruído técnico de mensagens retornadas por gateways/API."""
    if not value:
        return ""
    text = re.sub(r"<[^>]+>", " ", value)
    text = unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def is_rate_limit_message(value: Any) -> bool:
    text = str(value or "").upper()
    return "HTTP 429" in text or "TOO MANY REQUESTS" in text or "LIMITE TEMPORÁRIO DE REQUISIÇÕES" in text


def retry_after_seconds(headers: dict[str, str] | None, default_seconds: int, max_seconds: int) -> int:
    retry_after = None
    for key, value in (headers or {}).items():
        if str(key).lower() == "retry-after":
            retry_after = str(value).strip()
            break
    seconds: int | None = None
    if retry_after:
        if retry_after.isdigit():
            seconds = int(retry_after)
        else:
            try:
                retry_date = parsedate_to_datetime(retry_after)
                if retry_date.tzinfo is None:
                    retry_date = retry_date.replace(tzinfo=timezone.utc)
                seconds = int((retry_date - datetime.now(timezone.utc)).total_seconds())
            except (TypeError, ValueError, OverflowError):
                seconds = None
    seconds = default_seconds if seconds is None else seconds
    return max(1, min(int(seconds), int(max_seconds)))


def sleep_with_cancel(seconds: float, cancel_checker: Callable[[], bool] | None = None) -> bool:
    """Pausa a rotina sem perder a capacidade de cancelamento pelo usuário."""
    if seconds <= 0:
        return True
    end_at = time.monotonic() + float(seconds)
    while True:
        if cancel_checker and cancel_checker():
            return False
        remaining = end_at - time.monotonic()
        if remaining <= 0:
            return True
        time.sleep(min(1.0, remaining))


def http_error_message(nsu: int, status_code: int, preview: str | None = None) -> str:
    if status_code == 429:
        return (
            f"NSU {nsu}: limite temporário de requisições do Portal Nacional (HTTP 429). "
            "A rotina foi desacelerada automaticamente. Aguarde alguns minutos antes de uma nova tentativa caso o limite persista."
        )
    if status_code in {401, 403}:
        base = f"NSU {nsu}: HTTP {status_code}. Verifique autorização, certificado A1, procuração e permissões do CNPJ consultado."
    elif 500 <= status_code <= 599:
        base = f"NSU {nsu}: HTTP {status_code}. O Portal Nacional retornou indisponibilidade temporária."
    else:
        base = f"NSU {nsu}: HTTP {status_code}. A resposta não pôde ser processada."
    cleaned = clean_response_preview(preview)
    if cleaned:
        base += f" Prévia técnica: {cleaned}"
    return base


def municipios_from_notes(notes: list[dict[str, Any]]) -> str:
    municipios = sorted({str(n.get("municipio_prestacao") or "").strip() for n in notes if str(n.get("municipio_prestacao") or "").strip()})
    if not municipios:
        return "-"
    if len(municipios) <= 3:
        return ", ".join(municipios)
    return ", ".join(municipios[:3]) + f" +{len(municipios) - 3}"


def cliente_checkboxes(clientes: list[dict[str, Any]], selected_ids: list[int] | None = None, field_name: str = "cliente_ids") -> str:
    selected = set(normalize_id_list(selected_ids))
    if not clientes:
        return "<p class='muted small'>Nenhuma empresa cadastrada ainda.</p>"
    cards: list[str] = []
    for cliente in sorted(clientes, key=lambda c: c.get("razao_social", "")):
        cid = int(cliente.get("id"))
        checked = "checked" if cid in selected else ""
        cert_ok = "com certificado" if cliente.get("certificado_id") else "sem certificado"
        cards.append(
            "<label class='check-card'>"
            f"<input type='checkbox' name='{escape(field_name)}' value='{cid}' {checked}>"
            "<span class='check-dot'></span>"
            "<span>"
            f"<strong>{escape(cliente.get('razao_social',''))}</strong>"
            f"<em>{mask_cnpj(cliente.get('cnpj'))} · {escape(cert_ok)}</em>"
            "</span>"
            "</label>"
        )
    return "<div class='check-grid'>" + "".join(cards) + "</div>"


# ---------------------------------------------------------------------------
# Controle da consulta em andamento
# ---------------------------------------------------------------------------

ACTIVE_JOB_STATUSES = {"PENDENTE", "EM_ANDAMENTO", "PARANDO"}
FINAL_JOB_STATUSES = {"CONCLUIDA", "CONCLUIDA_COM_ERROS", "CANCELADA", "ERRO"}


JOB_NUMERIC_FIELDS = [
    "consultas_realizadas",
    "documentos_recebidos",
    "importadas",
    "atualizadas",
    "ignoradas",
    "fora_periodo",
]


def sync_status_kind(status: str | None) -> str:
    s = (status or "").upper()
    if s in {"CONCLUIDA", "OK", "AUTORIZADA"}:
        return "ok"
    if "LIMITE" in s or s in {"CONCLUIDA_COM_ERROS", "PARANDO", "CANCELADA", "AGUARDANDO", "EM CONSULTA", "INICIADA"}:
        return "warn"
    if "ERRO" in s or "SEM CERT" in s or "SEM_CERT" in s:
        return "err"
    return "info"


def latest_active_job(data: dict[str, Any] | None = None) -> dict[str, Any] | None:
    data = data or store.read()
    jobs = sorted(data.get("sync_jobs", []), key=lambda j: j.get("criado_em", ""), reverse=True)
    for job in jobs:
        if str(job.get("status") or "").upper() in ACTIVE_JOB_STATUSES:
            return job
    return None


def latest_job(data: dict[str, Any] | None = None) -> dict[str, Any] | None:
    data = data or store.read()
    jobs = sorted(data.get("sync_jobs", []), key=lambda j: j.get("criado_em", ""), reverse=True)
    return jobs[0] if jobs else None


def find_job(data: dict[str, Any], job_id: int | str | None) -> dict[str, Any] | None:
    if job_id in (None, ""):
        return None
    try:
        jid = int(job_id)
    except (TypeError, ValueError):
        return None
    for job in data.get("sync_jobs", []):
        if int(job.get("id", -1)) == jid:
            return job
    return None


def update_job(job_id: int, mutator: Callable[[dict[str, Any]], Any]) -> dict[str, Any] | None:
    data = store.read()
    job = find_job(data, job_id)
    if not job:
        return None
    mutator(job)
    job["atualizado_em"] = now_iso()
    store.write(data)
    return job


def job_recalculate_totals(job: dict[str, Any]) -> None:
    clientes = job.get("clientes", [])
    totals = {field: 0 for field in JOB_NUMERIC_FIELDS}
    for item in clientes:
        for field in JOB_NUMERIC_FIELDS:
            totals[field] += int(item.get(field) or 0)
    job["totais"] = totals
    job["empresas_processadas"] = sum(1 for item in clientes if str(item.get("status") or "").upper() in {"CONCLUIDA", "ERRO", "CANCELADA", "SEM CERTIFICADO"})
    job["empresas_com_erro"] = sum(1 for item in clientes if str(item.get("status") or "").upper() in {"ERRO", "SEM CERTIFICADO"})


def create_sync_job(
    clientes: list[dict[str, Any]],
    inicio: date | None,
    fim: date | None,
    max_consultas: int,
    parar_apos_vazios: int,
    reiniciar_nsu: bool,
    selected_ids: list[int] | None = None,
) -> dict[str, Any]:
    data = store.read()
    job_id = JsonStore.next_id(data, "job")
    job = {
        "id": job_id,
        "status": "PENDENTE",
        "criado_em": now_iso(),
        "atualizado_em": now_iso(),
        "finalizado_em": None,
        "cancelar_solicitado": False,
        "periodo_inicio": inicio.isoformat() if inicio else None,
        "periodo_fim": fim.isoformat() if fim else None,
        "parametros": {
            "max_consultas": int(max_consultas),
            "parar_apos_vazios": int(parar_apos_vazios),
            "reiniciar_nsu": bool(reiniciar_nsu),
            "cliente_ids": selected_ids or [],
        },
        "total_empresas": len(clientes),
        "empresas_processadas": 0,
        "empresas_com_erro": 0,
        "empresa_atual": None,
        "mensagem_atual": "Consulta criada. Aguardando início.",
        "totais": {field: 0 for field in JOB_NUMERIC_FIELDS},
        "clientes": [
            {
                "cliente_id": int(c.get("id")),
                "razao_social": c.get("razao_social", ""),
                "cnpj": c.get("cnpj", ""),
                "status": "AGUARDANDO",
                "nsu_inicial": int(c.get("ultimo_nsu") or 0) + 1,
                "ultimo_nsu_processado": int(c.get("ultimo_nsu") or 0),
                "consultas_realizadas": 0,
                "documentos_recebidos": 0,
                "importadas": 0,
                "atualizadas": 0,
                "ignoradas": 0,
                "fora_periodo": 0,
                "mensagens": [],
                "erro": None,
                "iniciado_em": None,
                "finalizado_em": None,
            }
            for c in clientes
        ],
    }
    data.setdefault("sync_jobs", []).append(job)
    # Mantém o arquivo leve; preserva os 50 jobs mais recentes.
    data["sync_jobs"] = sorted(data["sync_jobs"], key=lambda j: j.get("criado_em", ""), reverse=True)[:50]
    store.write(data)
    return job


def update_job_cliente_from_result(job_id: int, cliente_id: int, result: dict[str, Any], status: str | None = None) -> None:
    def mutate(job: dict[str, Any]) -> None:
        for item in job.get("clientes", []):
            if int(item.get("cliente_id", -1)) == int(cliente_id):
                if status:
                    item["status"] = status
                else:
                    item["status"] = result.get("status") or item.get("status") or "EM CONSULTA"
                for field in ["nsu_inicial", "ultimo_nsu_processado", *JOB_NUMERIC_FIELDS]:
                    if field in result:
                        item[field] = result.get(field)
                msgs = result.get("mensagens", []) or []
                item["mensagens"] = msgs[-20:]
                if item["status"] == "ERRO" or str(result.get("status") or "").upper() == "ERRO":
                    item["erro"] = msgs[-1] if msgs else "Erro na consulta."
                if str(item.get("status") or "").upper() in {"CONCLUIDA", "ERRO", "CANCELADA", "SEM CERTIFICADO"}:
                    item["finalizado_em"] = item.get("finalizado_em") or now_iso()
                break
        job_recalculate_totals(job)
    update_job(job_id, mutate)


def is_job_cancel_requested(job_id: int) -> bool:
    data = store.read()
    job = find_job(data, job_id)
    return bool(job and job.get("cancelar_solicitado"))


def request_stop_job(job_id: int) -> dict[str, Any] | None:
    def mutate(job: dict[str, Any]) -> None:
        if str(job.get("status") or "").upper() in ACTIVE_JOB_STATUSES:
            job["cancelar_solicitado"] = True
            job["status"] = "PARANDO"
            job["mensagem_atual"] = "Parada solicitada. A consulta será interrompida ao terminar o NSU em andamento."
    return update_job(job_id, mutate)


def run_sync_job(job_id: int) -> None:
    data = store.read()
    job = find_job(data, job_id)
    if not job:
        return
    params = job.get("parametros", {})
    inicio = parse_date_filter(job.get("periodo_inicio"))
    fim = parse_date_filter(job.get("periodo_fim"))
    max_consultas = int(params.get("max_consultas") or 500)
    parar_apos_vazios = int(params.get("parar_apos_vazios") or 3)
    reiniciar_nsu = bool(params.get("reiniciar_nsu"))

    def mark_job(status: str, message: str, current: dict[str, Any] | None = None, finished: bool = False) -> None:
        def mutate(j: dict[str, Any]) -> None:
            j["status"] = status
            j["mensagem_atual"] = message
            if current:
                j["empresa_atual"] = {"cliente_id": current.get("cliente_id"), "razao_social": current.get("razao_social"), "cnpj": current.get("cnpj")}
            if finished:
                j["finalizado_em"] = now_iso()
            job_recalculate_totals(j)
        update_job(job_id, mutate)

    mark_job("EM_ANDAMENTO", "Consulta iniciada.")
    any_errors = False
    canceled = False

    for item in list(job.get("clientes", [])):
        cliente_id = int(item.get("cliente_id"))
        data = store.read()
        cliente = find_by_id(data.get("clientes", []), cliente_id)
        if not cliente:
            any_errors = True
            update_job_cliente_from_result(job_id, cliente_id, sync_error_result({"id": cliente_id, "cnpj": item.get("cnpj"), "razao_social": item.get("razao_social")}, "Empresa não encontrada no cadastro.", inicio, fim), "ERRO")
            continue
        if is_job_cancel_requested(job_id):
            canceled = True
            update_job_cliente_from_result(job_id, cliente_id, sync_error_result(cliente, "Consulta cancelada antes de iniciar esta empresa.", inicio, fim), "CANCELADA")
            continue

        mark_job("EM_ANDAMENTO", f"Consultando {cliente.get('razao_social', '')}.", item)
        update_job_cliente_from_result(job_id, cliente_id, {"status": "EM CONSULTA", "mensagens": ["Iniciando consulta da empresa."], "nsu_inicial": (1 if reiniciar_nsu else int(cliente.get("ultimo_nsu") or 0) + 1)}, "EM CONSULTA")

        if not cliente.get("certificado_id"):
            any_errors = True
            result = sync_error_result(cliente, "Empresa sem certificado vinculado.", inicio, fim)
            update_job_cliente_from_result(job_id, cliente_id, result, "SEM CERTIFICADO")
            continue

        result = sincronizar_cliente(
            cliente_id,
            max_consultas=max_consultas,
            parar_apos_vazios=parar_apos_vazios,
            inicio=inicio,
            fim=fim,
            reiniciar_nsu=reiniciar_nsu,
            raise_on_error=False,
            cancel_checker=lambda jid=job_id: is_job_cancel_requested(jid),
            progress_callback=lambda res, cid=cliente_id: update_job_cliente_from_result(job_id, cid, res, res.get("status") or "EM CONSULTA"),
        )
        status = str(result.get("status") or "").upper()
        if status == "CANCELADA":
            canceled = True
        elif status == "ERRO":
            any_errors = True
        update_job_cliente_from_result(job_id, cliente_id, result, status if status else None)

    final_status = "CANCELADA" if canceled else "CONCLUIDA_COM_ERROS" if any_errors else "CONCLUIDA"
    final_message = "Consulta cancelada pelo usuário." if canceled else "Consulta finalizada com erros em uma ou mais empresas." if any_errors else "Consulta finalizada com sucesso."
    mark_job(final_status, final_message, finished=True)


def job_progress_percent(job: dict[str, Any]) -> int:
    total = int(job.get("total_empresas") or 0)
    if total <= 0:
        return 100 if str(job.get("status") or "").upper() in FINAL_JOB_STATUSES else 0
    done = int(job.get("empresas_processadas") or 0)
    current_bonus = 0.5 if job.get("empresa_atual") and str(job.get("status") or "").upper() in ACTIVE_JOB_STATUSES else 0
    return max(0, min(100, int(((done + current_bonus) / total) * 100)))


# ---------------------------------------------------------------------------
# HTML base
# ---------------------------------------------------------------------------


def render_page(title: str, active: str, body: str, subtitle: str | None = None) -> str:
    nav = [
        ("/", "dashboard", "Painel", "Visão geral"),
        ("/sincronizar", "sincronizar", "Busca", "Consulta ADN"),
        ("/clientes", "clientes", "Empresas", "CNPJ e certificado"),
        ("/retencoes", "retencoes", "Retenções", "Mapa fiscal"),
        ("/conferencia-excel", "conferencia", "Conferência", "Importar Excel"),
        ("/relatorio", "relatorio", "Notas", "Auditoria"),
        ("/sincronizacao/logs", "logs", "Histórico", "Logs"),
    ]
    active_label = next((label for _href, key, label, _desc in nav if key == active), title)
    nav_html = "".join(
        f"""
        <a class="side-nav-item {'active' if key == active else ''}" href="{href}">
          <span class="nav-glyph">{escape(label[:2].upper())}</span>
          <span class="nav-copy"><strong>{escape(label)}</strong><em>{escape(desc)}</em></span>
        </a>
        """
        for href, key, label, desc in nav
    )
    data = store.read()
    settings = get_settings()
    active_job = latest_active_job(data)
    if active_job:
        progress = job_progress_percent(active_job)
        consulta_controls = f"""
          <section class="job-widget active" aria-label="Consulta em andamento">
            <div class="job-widget-copy">
              <span>Busca em andamento</span>
              <strong>{progress}% concluído</strong>
              <div class="mini-progress"><i style="width:{progress}%"></i></div>
            </div>
            <div class="job-widget-actions">
              <a class="button blue compact-btn" href="/sincronizacao/progresso/{active_job.get('id')}">Acompanhar</a>
              <form method="post" action="/ui/sincronizacao/{active_job.get('id')}/parar" onsubmit="return confirm('Parar apenas a consulta de notas em andamento?');">
                <button class="stop compact-btn" type="submit">Pausar</button>
              </form>
            </div>
          </section>
        """
    else:
        consulta_controls = """
          <div class="top-actions">
            <a class="button blue compact-btn" href="/sincronizar">Nova busca</a>
            <a class="button ghost compact-btn" href="/retencoes">Retenções</a>
          </div>
        """
    sub = subtitle or "Gestão local de NFS-e, retenções e conferência fiscal."
    return f"""
    <!doctype html>
    <html lang="pt-BR">
      <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>{escape(title)} · ORA NFS-e</title>
        <link rel="icon" href="/static/favicon.png">
        <link rel="stylesheet" href="/static/ora.css?v=10">
      </head>
      <body>
        <div class="system-shell">
          <aside class="sidebar">
            <a class="brand-area" href="/" aria-label="ORA NFS-e">
              <img src="/static/ora_logo_white.png" alt="Grupo ORA">
              <span><strong>Fiscal Intelligence</strong><em>NFS-e Nacional</em></span>
            </a>
            <nav class="side-nav" aria-label="Navegação principal">
              {nav_html}
            </nav>
            <div class="sidebar-footer">
              <div class="local-badge">
                <span>Ambiente local</span>
                <strong>Dados em data/</strong>
              </div>
              <span class="version-chip">v{escape(str(settings.app_version))}</span>
            </div>
          </aside>

          <div class="main-area">
            <header class="topbar">
              <div class="page-heading">
                <div class="breadcrumb"><span>Grupo ORA</span><i>/</i><strong>{escape(active_label)}</strong></div>
                <h1>{escape(title)}</h1>
                <p>{escape(sub)}</p>
              </div>
              {consulta_controls}
            </header>

            <main class="workspace" id="conteudo">
              {body}
            </main>

            <footer class="footer-note small">
              <img src="/static/ora_logo_blue.png" alt="Grupo ORA" class="footer-logo">
              <span>Aplicação local. Certificados A1, senhas e XMLs fiscais permanecem sob controle do usuário.</span>
            </footer>
          </div>
        </div>
      </body>
    </html>
    """

# ---------------------------------------------------------------------------
# Dashboard e cadastros
# ---------------------------------------------------------------------------


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}



@app.get("/", response_class=HTMLResponse)
def home() -> str:
    data = store.read()
    notes = filtered_notes(data, papel="TODOS")
    prestado = totals_for_role(notes, "PRESTADOR")
    tomado = totals_for_role(notes, "TOMADOR")
    clientes = data.get("clientes", [])
    clientes_ativos = [c for c in clientes if c.get("ativo", True)]
    clientes_com_cert = sum(1 for c in clientes_ativos if c.get("certificado_id"))
    clientes_sem_cert = len(clientes_ativos) - clientes_com_cert
    ultimo_job = latest_job(data)

    total_servicos = prestado["valor_servico"] + tomado["valor_servico"]
    total_retencoes = prestado["total_retencoes"] + tomado["total_retencoes"]
    total_sociais = (
        prestado.get("valor_pis", Decimal("0")) + prestado.get("valor_cofins", Decimal("0")) + prestado.get("valor_csll", Decimal("0")) +
        tomado.get("valor_pis", Decimal("0")) + tomado.get("valor_cofins", Decimal("0")) + tomado.get("valor_csll", Decimal("0"))
    )

    metric_cards = f"""
      <div class="kpi-strip">
        <div class="card metric blue"><span class="metric-label">Movimento autorizado</span><strong class="metric-value">{fmt_money(total_servicos)}</strong><span class="metric-detail">prestadas + tomadas</span></div>
        <div class="card metric accent"><span class="metric-label">Total retido</span><strong class="metric-value">{fmt_money(total_retencoes)}</strong><span class="metric-detail">ISS + federais + outras</span></div>
        <div class="card metric"><span class="metric-label">PIS/COFINS/CSLL</span><strong class="metric-value">{fmt_money(total_sociais)}</strong><span class="metric-detail">retenção segregada</span></div>
        <div class="card metric dark"><span class="metric-label">Empresas aptas</span><strong class="metric-value">{clientes_com_cert}/{len(clientes_ativos)}</strong><span class="metric-detail">{clientes_sem_cert} sem certificado</span></div>
      </div>
    """

    status_job = "Sem busca recente"
    job_detail = "Execute uma busca para carregar NSUs e XMLs na base local."
    job_action = "/sincronizar"
    job_button = "Iniciar busca"
    job_kind = "info"
    if ultimo_job:
        status_job = str(ultimo_job.get("status") or "-")
        criado = str(ultimo_job.get("criado_em") or "").replace("T", " ")[:16]
        job_detail = f"{criado} · {ultimo_job.get('empresas_processadas', 0)} de {ultimo_job.get('total_empresas', 0)} empresas processadas"
        job_action = f"/sincronizacao/progresso/{ultimo_job.get('id')}"
        job_button = "Ver execução"
        job_kind = sync_status_kind(status_job)

    attention_items = []
    if clientes_sem_cert:
        attention_items.append(f"<li><strong>{clientes_sem_cert}</strong> empresa(s) ativa(s) sem certificado vinculado.</li>")
    if not notes:
        attention_items.append("<li>Nenhuma NFS-e importada. A primeira rotina recomendada é a busca de NFS-e.</li>")
    if total_sociais:
        attention_items.append("<li>Existem retenções sociais para revisão em PIS, COFINS e CSLL.</li>")
    if not attention_items:
        attention_items.append("<li>Base operacional sem pendências críticas de configuração.</li>")
    attention_html = "".join(attention_items)

    workflow = """
      <div class="module-list">
        <a class="module-row primary" href="/sincronizar"><span>01</span><div><strong>Executar busca</strong><em>Período, empresas, ritmo de consulta e acompanhamento.</em></div></a>
        <a class="module-row" href="/retencoes"><span>02</span><div><strong>Conferir retenções</strong><em>ISS, PIS, COFINS, CSLL, IRRF e INSS/CP separados.</em></div></a>
        <a class="module-row" href="/conferencia-excel"><span>03</span><div><strong>Comparar Excel</strong><em>Planilha externa contra a base puxada pelo sistema.</em></div></a>
        <a class="module-row" href="/relatorio"><span>04</span><div><strong>Auditar notas</strong><em>Detalhe por XML, prestador, tomador e critérios fiscais.</em></div></a>
      </div>
    """

    monthly = monthly_rows(notes)[:6]
    monthly_rows_html = "".join(
        f"<tr><td>{escape(row['mes'])}</td><td class='num'>{row['qtd_autorizadas']}</td><td class='num'>{fmt_money(row['faturamento_bruto'])}</td><td class='num'>{fmt_money(row['retencoes_faturamento'])}</td><td class='num'>{fmt_money(row['faturamento_liquido'])}</td></tr>"
        for row in monthly
    ) or "<tr><td colspan='5'>Sem movimento sincronizado ainda.</td></tr>"

    retention_tiles = f"""
      <div class="tax-ledger">
        <div><span>ISS retido</span><strong>{fmt_money(prestado['valor_iss_retido'] + tomado['valor_iss_retido'])}</strong></div>
        <div><span>PIS retido</span><strong>{fmt_money(prestado['valor_pis'] + tomado['valor_pis'])}</strong></div>
        <div><span>COFINS retido</span><strong>{fmt_money(prestado['valor_cofins'] + tomado['valor_cofins'])}</strong></div>
        <div><span>CSLL retida</span><strong>{fmt_money(prestado['valor_csll'] + tomado['valor_csll'])}</strong></div>
        <div><span>IRRF retido</span><strong>{fmt_money(prestado['valor_ir'] + tomado['valor_ir'])}</strong></div>
        <div><span>INSS/CP retido</span><strong>{fmt_money(prestado['valor_inss'] + tomado['valor_inss'])}</strong></div>
      </div>
    """

    body = f"""
      <section class="workbench-grid">
        <section class="card module-card">
          <div class="card-head">
            <div>
              <span class="small-eyebrow">Fluxo operacional</span>
              <h2>Rotina principal do sistema</h2>
            </div>
            <a class="button ghost compact-btn" href="/sincronizar">Abrir busca</a>
          </div>
          {workflow}
        </section>

        <aside class="card status-card">
          <div class="card-head">
            <div>
              <span class="small-eyebrow">Status da base</span>
              <h2>Próxima ação</h2>
            </div>
            {pill(status_job, job_kind)}
          </div>
          <p class="muted">{escape(job_detail)}</p>
          <a class="button blue" href="{job_action}">{job_button}</a>
          <div class="divider"></div>
          <ul class="attention-list">{attention_html}</ul>
        </aside>
      </section>

      {metric_cards}

      <section class="grid two">
        <section class="card data-card">
          <div class="card-head">
            <div><span class="small-eyebrow">Movimento mensal</span><h2>Últimos períodos sincronizados</h2></div>
            <a class="button ghost compact-btn" href="/relatorio">Detalhar</a>
          </div>
          <div class="table-wrap compact"><table><thead><tr><th>Mês</th><th class="num">Notas</th><th class="num">Serviços</th><th class="num">Retenções</th><th class="num">Líquido</th></tr></thead><tbody>{monthly_rows_html}</tbody></table></div>
        </section>

        <section class="card data-card">
          <div class="card-head">
            <div><span class="small-eyebrow">Mapa de retenções</span><h2>Tributos retidos identificados</h2></div>
            <a class="button ghost compact-btn" href="/retencoes">Conferir</a>
          </div>
          {retention_tiles}
        </section>
      </section>

      <section class="notice system-note">
        <strong>Leitura fiscal aplicada</strong>
        <span>O sistema mantém PIS/COFINS apurados no XML separados das retenções na fonte. Retenções sociais são exibidas por PIS, COFINS e CSLL quando houver indicação no XML.</span>
      </section>
    """
    return render_page("Painel operacional", "dashboard", body, subtitle="Visão de trabalho para busca, conferência e auditoria de NFS-e.")

def render_certificados_table(data: dict[str, Any]) -> str:
    certificados = data.get("certificados", [])
    cert_rows = "".join(
        "<tr>"
        f"<td>{cert.get('id')}</td>"
        f"<td><strong>{escape(cert.get('alias',''))}</strong><br><span class='small muted'>{escape(cert.get('nome_original',''))}</span></td>"
        f"<td>{escape(cert.get('cnpj_raiz') or '')}</td>"
        f"<td>{pill('ativo' if cert.get('ativo', True) else 'inativo', 'ok' if cert.get('ativo', True) else 'warn')}</td>"
        "<td class='nowrap'>"
        f"<form method='post' action='/ui/certificados/{cert.get('id')}/excluir' onsubmit=\"return confirm('Excluir este certificado? Os clientes vinculados ficarão sem certificado até novo vínculo.');\" style='display:inline'>"
        "<button class='ghost danger-button' type='submit'>Excluir</button></form>"
        "</td>"
        "</tr>"
        for cert in certificados
    ) or "<tr><td colspan='5'>Nenhum certificado cadastrado.</td></tr>"
    return f"<div class='table-wrap'><table><thead><tr><th>ID</th><th>Certificado</th><th>Raiz CNPJ</th><th>Status</th><th>Ações</th></tr></thead><tbody>{cert_rows}</tbody></table></div>"


@app.get("/certificados")
def certificados_page() -> RedirectResponse:
    return RedirectResponse("/clientes#certificados", status_code=303)


def render_clientes_table(data: dict[str, Any]) -> str:
    clientes = data.get("clientes", [])
    certificados = data.get("certificados", [])
    cert_by_id = {int(c["id"]): c for c in certificados if "id" in c}
    cliente_rows = []
    for cliente in sorted(clientes, key=lambda c: c.get("razao_social", "")):
        cert = cert_by_id.get(int(cliente.get("certificado_id") or -1))
        last_sync = cliente.get("ultima_sincronizacao_em") or "nunca"
        if last_sync != "nunca":
            last_sync = last_sync.replace("T", " ")[:16]
        cliente_rows.append(
            "<tr>"
            f"<td>{cliente.get('id')}</td>"
            f"<td><strong>{escape(cliente.get('razao_social',''))}</strong></td>"
            f"<td>{mask_cnpj(cliente.get('cnpj'))}</td>"
            f"<td>{escape(cert.get('alias','') if cert else 'sem certificado')}</td>"
            f"<td class='num'>{cliente.get('ultimo_nsu', 0)}</td>"
            f"<td class='nowrap'>{escape(last_sync)}</td>"
            f"<td>{pill('ativo' if cliente.get('ativo', True) else 'inativo', 'ok' if cliente.get('ativo', True) else 'warn')}</td>"
            "<td class='nowrap actions-cell'>"
            f"<a class='button blue compact-btn' href='/sincronizar?cliente_ids={cliente.get('id')}'>Buscar</a> "
            f"<a class='button ghost compact-btn' href='/relatorio?cliente_id={cliente.get('id')}&papel=TODOS'>Relatório</a> "
            f"<form method='post' action='/ui/clientes/{cliente.get('id')}/excluir' onsubmit=\"return confirm('Excluir esta empresa e as notas/logs vinculados a ela?');\" style='display:inline'>"
            "<button class='ghost danger-button compact-btn' type='submit'>Excluir</button></form>"
            "</td>"
            "</tr>"
        )
    body = "".join(cliente_rows) or "<tr><td colspan='8'>Nenhuma empresa cadastrada.</td></tr>"
    return f"<div class='table-wrap'><table><thead><tr><th>ID</th><th>Empresa</th><th>CNPJ</th><th>Certificado</th><th class='num'>Último NSU</th><th>Última busca</th><th>Status</th><th>Ações</th></tr></thead><tbody>{body}</tbody></table></div>"


@app.get("/clientes", response_class=HTMLResponse)
def clientes_page() -> str:
    data = store.read()
    certificados = data.get("certificados", [])
    clientes = data.get("clientes", [])
    cert_options = [("", "Selecionar depois")] + [(c["id"], f"{c.get('alias','')} · {'ativo' if c.get('ativo', True) else 'inativo'}") for c in certificados]
    clientes_ativos = [c for c in clientes if c.get("ativo", True)]
    prontas = sum(1 for c in clientes_ativos if c.get("certificado_id"))

    body = f"""
      <section class="toolbar-card">
        <div class="toolbar-copy">
          <span class="small-eyebrow">Cadastro operacional</span>
          <h2>Empresas e certificados</h2>
          <p>Cadastre CNPJs, vincule certificados A1 e acompanhe quais empresas estão aptas para consulta.</p>
        </div>
        <div class="toolbar-stats">
          <div><span>Empresas ativas</span><strong>{len(clientes_ativos)}</strong></div>
          <div><span>Aptas para busca</span><strong>{prontas}</strong></div>
          <div><span>Certificados</span><strong>{len(certificados)}</strong></div>
        </div>
      </section>

      <section class="grid two anchor-grid" id="certificados">
        <section class="card form-card">
          <div class="card-head">
            <div><span class="small-eyebrow">Certificado A1</span><h2>Adicionar certificado</h2></div>
          </div>
          <p class="muted small">O arquivo e a senha ficam locais. Use um apelido claro para identificar o vínculo com as empresas.</p>
          <form method="post" action="/ui/certificados" enctype="multipart/form-data">
            <label>Apelido do certificado</label>
            <input name="alias" placeholder="Ex.: Cliente X — A1 2026" required>
            <label>Arquivo .pfx ou .p12</label>
            <input type="file" name="arquivo" accept=".pfx,.p12" required>
            <label>Senha</label>
            <input type="password" name="senha" placeholder="senha do certificado">
            <p class="small warning">Não envie certificado ou senha por canais externos.</p>
            <button type="submit">Salvar certificado</button>
          </form>
        </section>

        <section class="card form-card">
          <div class="card-head">
            <div><span class="small-eyebrow">Empresa</span><h2>Adicionar empresa</h2></div>
          </div>
          <p class="muted small">Cadastre o CNPJ e vincule o certificado autorizado. O município é apenas informativo.</p>
          <form method="post" action="/ui/clientes">
            <label>Razão social</label>
            <input name="razao_social" required>
            <label>CNPJ</label>
            <input name="cnpj" placeholder="00.000.000/0001-00" required>
            <label>Certificado</label>
            <select name="certificado_id">{html_select_options(cert_options, '')}</select>
            <button type="submit">Cadastrar empresa</button>
          </form>
        </section>
      </section>

      <section class="card data-card">
        <div class="card-head">
          <div><span class="small-eyebrow">Base de empresas</span><h2>Empresas cadastradas</h2></div>
          <a class="button blue compact-btn" href="/sincronizar">Buscar NFS-e</a>
        </div>
        {render_clientes_table(data)}
      </section>

      <section class="card data-card">
        <div class="card-head">
          <div><span class="small-eyebrow">Credenciais locais</span><h2>Certificados cadastrados</h2></div>
          <span class="muted small">Exclua apenas certificados sem autorização vigente.</span>
        </div>
        {render_certificados_table(data)}
      </section>
    """
    return render_page("Empresas", "clientes", body, subtitle="Cadastro de CNPJs, certificados e prontidão para consulta.")

@app.get("/sincronizar", response_class=HTMLResponse)
def sincronizar_page(
    cliente_ids: list[int] = Query(default=[]),
    inicio: str | None = Query(default=None),
    fim: str | None = Query(default=None),
) -> str:
    data = store.read()
    settings = get_settings()
    clientes = [c for c in data.get("clientes", []) if c.get("ativo", True)]
    selected = normalize_id_list(cliente_ids)
    ready = sum(1 for c in clientes if c.get("certificado_id"))
    active_job = latest_active_job(data)

    active_job_html = ""
    if active_job:
        active_job_html = f"""
          <section class='notice active-job-notice'>
            <div>
              <strong>Existe uma busca em andamento.</strong>
              <span>Evite iniciar outra consulta ao mesmo tempo para reduzir duplicidade de NSU e limite HTTP 429.</span>
            </div>
            <div class='actions'>
              <a class='button blue compact-btn' href='/sincronizacao/progresso/{active_job.get('id')}'>Acompanhar</a>
              <form method='post' action='/ui/sincronizacao/{active_job.get('id')}/parar' onsubmit="return confirm('Parar apenas a consulta de notas em andamento?');"><button class='stop compact-btn' type='submit'>Pausar</button></form>
            </div>
          </section>
        """

    clientes_table = "".join(
        "<tr>"
        f"<td><strong>{escape(c.get('razao_social',''))}</strong><br><span class='small muted'>{mask_cnpj(c.get('cnpj'))}</span></td>"
        f"<td>{pill('apta' if c.get('certificado_id') else 'sem certificado', 'ok' if c.get('certificado_id') else 'warn')}</td>"
        f"<td class='num'>{c.get('ultimo_nsu', 0)}</td>"
        f"<td><a class='button ghost compact-btn' href='/relatorio?cliente_id={c.get('id')}&papel=TODOS'>Notas</a></td>"
        "</tr>"
        for c in sorted(clientes, key=lambda x: x.get("razao_social", ""))
    ) or "<tr><td colspan='4'>Nenhuma empresa cadastrada.</td></tr>"

    body = f"""
      {active_job_html}

      <section class="ops-layout">
        <section class="card command-form-card">
          <div class="card-head">
            <div>
              <span class="small-eyebrow">Consulta ADN</span>
              <h2>Nova busca de NFS-e</h2>
            </div>
            <div class="inline-status">
              <span>{ready} aptas</span>
              <span>{len(clientes)} ativas</span>
            </div>
          </div>

          <form method="post" action="/ui/sincronizar-todos">
            <div class="form-row two-cols">
              <div><label>Início do período</label><input type="date" name="inicio" value="{escape(inicio or '')}"></div>
              <div><label>Fim do período</label><input type="date" name="fim" value="{escape(fim or '')}"></div>
            </div>

            <div class="context-line compact-context">
              <div><span>Período</span><strong>{escape(periodo_label(inicio, fim))}</strong></div>
              <div><span>Escopo</span><strong>{'selecionadas' if selected else 'todas as aptas'}</strong></div>
              <div><span>Limite 429</span><strong>pausa automática</strong></div>
            </div>

            <div class="field-block">
              <label>Empresas para consultar</label>
              <p class="muted small">Sem seleção, o sistema consulta todas as empresas ativas com certificado vinculado.</p>
              {cliente_checkboxes(clientes, selected)}
            </div>

            <details class="advanced-box">
              <summary>Parâmetros avançados de NSU</summary>
              <div class="form-row three-cols">
                <div><label>Máximo de NSUs por empresa</label><input name="max_consultas" value="500"></div>
                <div><label>Parar após NSUs vazios</label><input name="parar_apos_vazios" value="3"></div>
                <label class="switch-line"><input type="checkbox" name="reiniciar_nsu" value="1"><span>Reconsultar histórico desde o NSU 1</span></label>
              </div>
              <div class="notice rate-limit-note">
                <strong>Ritmo de consulta</strong>
                <span>Intervalo: {settings.request_delay_seconds}s · pausa 429: {settings.rate_limit_pause_seconds}s · retentativas: {settings.max_rate_limit_retries}.</span>
              </div>
            </details>

            <div class="actions command-actions">
              <button class="blue" type="submit">Iniciar busca</button>
              <a class="button ghost" href="/clientes">Ajustar empresas</a>
            </div>
          </form>
        </section>

        <aside class="card ops-aside">
          <div class="card-head">
            <div><span class="small-eyebrow">Checklist</span><h2>Antes de executar</h2></div>
          </div>
          <div class="system-checklist">
            <div><strong>1</strong><span>Período preenchido quando a conferência exigir recorte.</span></div>
            <div><strong>2</strong><span>Empresas com certificado A1 vinculado.</span></div>
            <div><strong>3</strong><span>Buscar uma rotina por vez para evitar HTTP 429.</span></div>
          </div>
          <div class="divider"></div>
          <div class="mini-metrics">
            <span><strong>{settings.request_delay_seconds}s</strong> intervalo</span>
            <span><strong>{settings.rate_limit_pause_seconds}s</strong> pausa 429</span>
            <span><strong>{settings.max_rate_limit_retries}</strong> retentativas</span>
          </div>
        </aside>
      </section>

      <section class="card data-card">
        <div class="card-head">
          <div><span class="small-eyebrow">Prontidão</span><h2>Empresas disponíveis para consulta</h2></div>
          <span class="muted small">Certificado e último NSU antes da execução.</span>
        </div>
        <div class="table-wrap"><table><thead><tr><th>Empresa</th><th>Status</th><th class="num">Último NSU</th><th>Atalho</th></tr></thead><tbody>{clientes_table}</tbody></table></div>
      </section>
    """
    return render_page("Busca de NFS-e", "sincronizar", body, subtitle="Execução operacional com controle de NSU, empresas e ritmo de consulta.")

@app.post("/ui/certificados")
async def ui_criar_certificado(alias: str = Form(...), senha: str = Form(default=""), arquivo: UploadFile = File(...)) -> RedirectResponse:
    await criar_certificado(alias, senha or None, arquivo)
    return RedirectResponse("/clientes#certificados", status_code=303)


@app.post("/ui/clientes")
def ui_criar_cliente(
    razao_social: str = Form(...),
    cnpj: str = Form(...),
    municipio: str = Form(default="Portal Nacional"),
    certificado_id: str = Form(default=""),
) -> RedirectResponse:
    cert_id = int(certificado_id) if str(certificado_id).strip() else None
    criar_cliente(razao_social=razao_social, cnpj=cnpj, municipio=municipio, certificado_id=cert_id)
    return RedirectResponse("/clientes", status_code=303)


@app.post("/ui/clientes/{cliente_id}/excluir")
def ui_excluir_cliente(cliente_id: int) -> RedirectResponse:
    excluir_cliente(cliente_id)
    return RedirectResponse("/clientes", status_code=303)


@app.post("/ui/certificados/{certificado_id}/excluir")
def ui_excluir_certificado(certificado_id: int) -> RedirectResponse:
    excluir_certificado(certificado_id)
    return RedirectResponse("/clientes#certificados", status_code=303)


def excluir_cliente(cliente_id: int) -> dict[str, Any]:
    data = store.read()
    cliente = find_by_id(data.get("clientes", []), cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Empresa não encontrada.")
    cnpj_cliente = digits(cliente.get("cnpj"))
    data["clientes"] = [c for c in data.get("clientes", []) if int(c.get("id", -1)) != int(cliente_id)]
    data["notas"] = [n for n in data.get("notas", []) if int(n.get("cliente_id") or -1) != int(cliente_id) and digits(n.get("cnpj_cliente")) != cnpj_cliente]
    data["logs"] = [l for l in data.get("logs", []) if int(l.get("cliente_id") or -1) != int(cliente_id) and digits(l.get("cnpj_cliente")) != cnpj_cliente]
    for job in data.get("sync_jobs", []):
        job["clientes"] = [i for i in job.get("clientes", []) if int(i.get("cliente_id") or -1) != int(cliente_id)]
        job_recalculate_totals(job)
    store.write(data)
    return {"status": "excluida", "cliente_id": cliente_id}


def excluir_certificado(certificado_id: int) -> dict[str, Any]:
    data = store.read()
    certificados = data.get("certificados", [])
    certificado = find_by_id(certificados, certificado_id)
    if not certificado:
        raise HTTPException(status_code=404, detail="Certificado não encontrado.")

    # Remove o arquivo físico do A1 e desvincula clientes.
    caminho = certificado.get("caminho_arquivo")
    if caminho:
        try:
            Path(caminho).unlink(missing_ok=True)
        except OSError:
            pass
    for cliente in data.get("clientes", []):
        if int(cliente.get("certificado_id") or -1) == int(certificado_id):
            cliente["certificado_id"] = None
            cliente["observacao_certificado"] = f"Certificado {certificado_id} excluído em {now_iso()}"
    data["certificados"] = [c for c in certificados if int(c.get("id", -1)) != int(certificado_id)]
    store.write(data)
    return {"status": "excluido", "certificado_id": certificado_id}


@app.post("/ui/parar-servico", response_class=HTMLResponse)
def ui_parar_servico() -> str:
    """Compatibilidade com versões antigas: agora o botão para somente a consulta de notas."""
    data = store.read()
    active = latest_active_job(data)
    if active:
        request_stop_job(int(active["id"]))
        return RedirectResponse(f"/sincronizacao/progresso/{active['id']}", status_code=303)
    body = """
      <section class="card dark center-card">
        <div class="ora-glyph"></div>
        <h2>Nenhuma consulta em andamento</h2>
        <p class="muted small">O sistema continua aberto. Para iniciar uma nova busca, acesse a etapa <span class="codeish">3 · Buscar NFS-e</span>.</p>
        <div class="actions" style="justify-content:center; margin-top:18px;"><a class="button blue" href="/sincronizar">Ir para busca</a></div>
      </section>
    """
    return render_page("Parar consulta", "sincronizar", body, subtitle="O controle agora interrompe somente a consulta das notas, não o sistema local.")


async def criar_certificado(alias: str, senha: str | None, arquivo: UploadFile) -> dict[str, Any]:
    data = store.read()
    if any(c.get("alias", "").lower() == alias.strip().lower() for c in data.get("certificados", [])):
        raise HTTPException(status_code=409, detail="Já existe certificado com esse alias.")
    filename = arquivo.filename or "certificado.pfx"
    if not filename.lower().endswith((".pfx", ".p12")):
        raise HTTPException(status_code=400, detail="Envie um certificado A1 .pfx ou .p12.")

    settings = get_settings()
    cert_dir = Path(settings.cert_store_dir)
    cert_dir.mkdir(parents=True, exist_ok=True)
    cert_id = JsonStore.next_id(data, "certificado")
    safe = f"{cert_id:04d}_{sanitize_filename(alias)}{Path(filename).suffix.lower()}"
    path = cert_dir / safe
    content = await arquivo.read()
    path.write_bytes(content)

    try:
        cnpj_raiz = extract_cnpj_root_from_pfx(path, senha)
    except CertificadoErro as exc:
        path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    cert = {
        "id": cert_id,
        "alias": alias.strip(),
        "cnpj_raiz": cnpj_raiz,
        "caminho_arquivo": str(path),
        "nome_original": filename,
        "senha_criptografada": encrypt_secret(senha or ""),
        "ativo": True,
        "criado_em": now_iso(),
    }
    data["certificados"].append(cert)
    store.write(data)
    return cert


def criar_cliente(razao_social: str, cnpj: str, municipio: str, certificado_id: int | None) -> dict[str, Any]:
    data = store.read()
    cnpj_digits = digits(cnpj)
    if not cnpj_digits or len(cnpj_digits) != 14:
        raise HTTPException(status_code=400, detail="Informe um CNPJ válido com 14 dígitos.")
    if find_cliente_by_cnpj(data, cnpj_digits):
        raise HTTPException(status_code=409, detail="Cliente já cadastrado.")
    if certificado_id and not find_by_id(data.get("certificados", []), certificado_id):
        raise HTTPException(status_code=404, detail="Certificado não encontrado.")
    cliente = {
        "id": JsonStore.next_id(data, "cliente"),
        "razao_social": razao_social.strip(),
        "cnpj": cnpj_digits,
        "municipio": municipio.strip() or "Portal Nacional",
        "certificado_id": certificado_id,
        "ultimo_nsu": 0,
        "ativo": True,
        "criado_em": now_iso(),
        "ultima_sincronizacao_em": None,
    }
    data["clientes"].append(cliente)
    store.write(data)
    return cliente


# ---------------------------------------------------------------------------
# Processamento de XMLs e sincronização ADN
# ---------------------------------------------------------------------------


def save_xml_file(cnpj_cliente: str, parsed: ParsedNfse) -> str:
    settings = get_settings()
    folder = Path(settings.xml_store_dir) / (digits(cnpj_cliente) or "sem_cnpj")
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{parsed.xml_hash}.xml"
    if not path.exists():
        path.write_text(parsed.xml_original, encoding="utf-8")
    return str(path)


def note_dict_from_parsed(cliente: dict[str, Any], parsed: ParsedNfse, origem: str, nsu: int | None = None, label: str | None = None) -> dict[str, Any] | None:
    papel = define_papel_cliente(cliente["cnpj"], parsed)
    if papel == "DESCONHECIDO":
        return None
    values = parsed.to_jsonable()
    values.pop("xml_original", None)
    values["id"] = f"{digits(cliente['cnpj'])}:{parsed.chave_acesso}"
    values["cnpj_cliente"] = digits(cliente["cnpj"])
    values["cliente_id"] = cliente["id"]
    values["cliente_razao_social"] = cliente.get("razao_social")
    values["papel_cliente"] = papel
    values["origem"] = origem
    values["nsu"] = nsu
    values["label_origem"] = label
    values["xml_path"] = save_xml_file(cliente["cnpj"], parsed)
    values["atualizado_em"] = now_iso()
    return values


def upsert_note(data: dict[str, Any], note: dict[str, Any]) -> str:
    notes = data.setdefault("notas", [])
    for index, existing in enumerate(notes):
        if existing.get("id") == note.get("id"):
            created = existing.get("criado_em") or now_iso()
            merged = {**existing, **note, "criado_em": created, "atualizado_em": now_iso()}
            if merged != existing:
                notes[index] = merged
                return "atualizada"
            return "ignorada"
    note["criado_em"] = now_iso()
    notes.append(note)
    return "importada"


def process_xml_items(
    data: dict[str, Any],
    cliente: dict[str, Any],
    xml_items: list[tuple[str, bytes]],
    origem: str,
    nsu: int | None = None,
    inicio: date | None = None,
    fim: date | None = None,
) -> dict[str, Any]:
    counts = {
        "notas_lidas": 0,
        "importadas": 0,
        "atualizadas": 0,
        "ignoradas": 0,
        "fora_periodo": 0,
        "mensagens": [],
    }
    for label, xml_bytes in xml_items:
        try:
            parsed_notes = parse_nfses_from_xml_bytes(xml_bytes)
        except Exception as exc:  # noqa: BLE001
            counts["mensagens"].append(f"{label}: XML não lido como NFS-e: {exc}")
            continue
        counts["notas_lidas"] += len(parsed_notes)
        for parsed in parsed_notes:
            note = note_dict_from_parsed(cliente, parsed, origem=origem, nsu=nsu, label=label)
            if note is None:
                counts["ignoradas"] += 1
                continue
            note_date = parse_date_filter(str(note.get("competencia") or note.get("data_emissao") or ""))
            if inicio and note_date and note_date < inicio:
                counts["fora_periodo"] += 1
                continue
            if fim and note_date and note_date > fim:
                counts["fora_periodo"] += 1
                continue
            action = upsert_note(data, note)
            if action == "importada":
                counts["importadas"] += 1
            elif action == "atualizada":
                counts["atualizadas"] += 1
            else:
                counts["ignoradas"] += 1
    return counts


def reprocess_saved_retencoes(data: dict[str, Any]) -> dict[str, Any]:
    """Recalcula notas já gravadas usando o XML armazenado localmente.

    Útil após ajustes de parser fiscal, sem depender de nova consulta ao ADN.
    """
    counts = {
        "avaliadas": 0,
        "atualizadas": 0,
        "sem_xml": 0,
        "falhas": 0,
        "mensagens": [],
    }
    notas = data.setdefault("notas", [])
    clientes = data.get("clientes", [])

    for index, existing in enumerate(list(notas)):
        counts["avaliadas"] += 1
        xml_path = existing.get("xml_path")
        if not xml_path:
            counts["sem_xml"] += 1
            continue

        path = Path(str(xml_path))
        if not path.exists():
            counts["sem_xml"] += 1
            continue

        cliente = find_by_id(clientes, existing.get("cliente_id"))
        if not cliente and existing.get("cnpj_cliente"):
            cliente = find_cliente_by_cnpj(data, str(existing.get("cnpj_cliente")))
        if not cliente:
            counts["falhas"] += 1
            if len(counts["mensagens"]) < 8:
                counts["mensagens"].append(f"Nota {existing.get('numero') or existing.get('chave_acesso')}: empresa não localizada.")
            continue

        try:
            parsed_notes = parse_nfses_from_xml_bytes(path.read_bytes())
        except Exception as exc:  # noqa: BLE001
            counts["falhas"] += 1
            if len(counts["mensagens"]) < 8:
                counts["mensagens"].append(f"{path.name}: XML não reprocessado: {exc}")
            continue

        refreshed: dict[str, Any] | None = None
        for parsed in parsed_notes:
            candidate = note_dict_from_parsed(
                cliente,
                parsed,
                origem=str(existing.get("origem") or "REPROCESSAMENTO_XML"),
                nsu=existing.get("nsu"),
                label=str(existing.get("label_origem") or path.name),
            )
            if not candidate:
                continue
            if candidate.get("id") == existing.get("id") or candidate.get("chave_acesso") == existing.get("chave_acesso"):
                refreshed = candidate
                break

        if not refreshed:
            counts["falhas"] += 1
            if len(counts["mensagens"]) < 8:
                counts["mensagens"].append(f"{path.name}: NFS-e correspondente não localizada no XML.")
            continue

        updated_at = now_iso()
        merged = {
            **existing,
            **refreshed,
            "criado_em": existing.get("criado_em") or updated_at,
            "atualizado_em": updated_at,
            "reprocessado_em": updated_at,
        }
        # Mantém rastreabilidade da origem original.
        if existing.get("origem"):
            merged["origem"] = existing.get("origem")
        if existing.get("nsu") is not None:
            merged["nsu"] = existing.get("nsu")
        if existing.get("label_origem"):
            merged["label_origem"] = existing.get("label_origem")

        notas[index] = merged
        counts["atualizadas"] += 1

    data.setdefault("logs", []).append({
        "id": JsonStore.next_id(data, "log"),
        "origem": "REPROCESSAMENTO_RETENCOES",
        "status": "CONCLUIDA" if counts["falhas"] == 0 else "CONCLUIDA_COM_ALERTAS",
        "criado_em": now_iso(),
        "finalizado_em": now_iso(),
        "mensagens": [
            f"{counts['atualizadas']} nota(s) recalculada(s) a partir dos XMLs salvos.",
            f"{counts['sem_xml']} nota(s) sem XML local para reprocessamento.",
            f"{counts['falhas']} falha(s) de reprocessamento.",
            *counts["mensagens"],
        ],
    })
    return counts


@app.post("/ui/reprocessar-retencoes")
def ui_reprocessar_retencoes() -> RedirectResponse:
    result = store.transaction(reprocess_saved_retencoes)
    query = urlencode({
        "reprocessadas": result.get("atualizadas", 0),
        "sem_xml": result.get("sem_xml", 0),
        "falhas": result.get("falhas", 0),
    })
    return RedirectResponse(f"/retencoes?{query}", status_code=303)


@app.post("/ui/clientes/{cliente_id}/sincronizar")
def ui_sincronizar_cliente(
    cliente_id: int,
    background_tasks: BackgroundTasks,
    max_consultas: int = Form(default=500),
    parar_apos_vazios: int = Form(default=3),
    inicio: str | None = Form(default=None),
    fim: str | None = Form(default=None),
    reiniciar_nsu: str | None = Form(default=None),
) -> RedirectResponse:
    data = store.read()
    active = latest_active_job(data)
    if active:
        return RedirectResponse(f"/sincronizacao/progresso/{active['id']}", status_code=303)
    cliente = find_by_id(data.get("clientes", []), cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Empresa não encontrada.")
    inicio_d = parse_date_filter(inicio)
    fim_d = parse_date_filter(fim)
    job = create_sync_job([cliente], inicio_d, fim_d, max_consultas, parar_apos_vazios, bool(reiniciar_nsu), [cliente_id])
    background_tasks.add_task(run_sync_job, int(job["id"]))
    return RedirectResponse(f"/sincronizacao/progresso/{job['id']}", status_code=303)


def sync_error_result(cliente: dict[str, Any], message: str, inicio: date | None, fim: date | None) -> dict[str, Any]:
    return {
        "id": None,
        "cliente_id": cliente.get("id"),
        "cliente_razao_social": cliente.get("razao_social"),
        "cnpj_cliente": cliente.get("cnpj"),
        "origem": "ADN_API",
        "status": "ERRO",
        "periodo_inicio": inicio.isoformat() if inicio else None,
        "periodo_fim": fim.isoformat() if fim else None,
        "nsu_inicial": int(cliente.get("ultimo_nsu") or 0) + 1,
        "ultimo_nsu_processado": int(cliente.get("ultimo_nsu") or 0),
        "consultas_realizadas": 0,
        "documentos_recebidos": 0,
        "importadas": 0,
        "atualizadas": 0,
        "ignoradas": 0,
        "fora_periodo": 0,
        "mensagens": [message],
        "erro": message,
        "criado_em": now_iso(),
        "finalizado_em": now_iso(),
    }


@app.post("/ui/sincronizar-todos")
def ui_sincronizar_todos(
    background_tasks: BackgroundTasks,
    max_consultas: int = Form(default=500),
    parar_apos_vazios: int = Form(default=3),
    inicio: str | None = Form(default=None),
    fim: str | None = Form(default=None),
    cliente_ids: list[int] | None = Form(default=None),
    reiniciar_nsu: str | None = Form(default=None),
) -> RedirectResponse:
    data = store.read()
    active = latest_active_job(data)
    if active:
        return RedirectResponse(f"/sincronizacao/progresso/{active['id']}", status_code=303)
    selected_ids = set(normalize_id_list(cliente_ids))
    inicio_d = parse_date_filter(inicio)
    fim_d = parse_date_filter(fim)
    clientes = [c for c in data.get("clientes", []) if c.get("ativo", True) and (not selected_ids or int(c.get("id")) in selected_ids)]
    job = create_sync_job(clientes, inicio_d, fim_d, max_consultas, parar_apos_vazios, bool(reiniciar_nsu), list(selected_ids))
    background_tasks.add_task(run_sync_job, int(job["id"]))
    return RedirectResponse(f"/sincronizacao/progresso/{job['id']}", status_code=303)


def sincronizar_cliente(
    cliente_id: int,
    max_consultas: int = 500,
    parar_apos_vazios: int = 3,
    inicio: date | None = None,
    fim: date | None = None,
    reiniciar_nsu: bool = False,
    raise_on_error: bool = True,
    cancel_checker: Callable[[], bool] | None = None,
    progress_callback: Callable[[dict[str, Any]], None] | None = None,
) -> dict[str, Any]:
    settings = get_settings()
    data = store.read()
    cliente = find_by_id(data.get("clientes", []), cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado.")
    certificado = find_by_id(data.get("certificados", []), cliente.get("certificado_id"))
    if not certificado or not certificado.get("ativo", True):
        if raise_on_error:
            raise HTTPException(status_code=400, detail="Cliente sem certificado ativo vinculado.")
        return sync_error_result(cliente, "Cliente sem certificado ativo vinculado.", inicio, fim)

    nsu_start = 1 if reiniciar_nsu else int(cliente.get("ultimo_nsu") or 0) + 1
    log = {
        "id": JsonStore.next_id(data, "log"),
        "cliente_id": cliente["id"],
        "cliente_razao_social": cliente.get("razao_social"),
        "cnpj_cliente": cliente["cnpj"],
        "origem": "ADN_API",
        "status": "INICIADA",
        "periodo_inicio": inicio.isoformat() if inicio else None,
        "periodo_fim": fim.isoformat() if fim else None,
        "reiniciar_nsu": bool(reiniciar_nsu),
        "nsu_inicial": nsu_start,
        "ultimo_nsu_processado": int(cliente.get("ultimo_nsu") or 0),
        "consultas_realizadas": 0,
        "documentos_recebidos": 0,
        "importadas": 0,
        "atualizadas": 0,
        "ignoradas": 0,
        "fora_periodo": 0,
        "mensagens": [],
        "criado_em": now_iso(),
        "finalizado_em": None,
    }
    data.setdefault("logs", []).append(log)
    store.write(data)

    result = dict(log)
    if progress_callback:
        progress_callback({**result, "status": "EM CONSULTA", "mensagens": ["Preparando certificado e conexão com o ADN."]})
    try:
        senha = decrypt_secret(certificado.get("senha_criptografada"))
        client = AdnClient(settings.nfse_adn_base_url, certificado["caminho_arquivo"], senha, timeout=settings.request_timeout_seconds)
    except (SecretStoreError, AdnClientError, KeyError) as exc:
        result["status"] = "ERRO"
        result["mensagens"].append(str(exc))
        if progress_callback:
            progress_callback(result)
        _finish_log(result)
        if raise_on_error:
            raise HTTPException(status_code=500, detail=str(exc)) from exc
        return result

    empty_streak = 0
    status_final = "CONCLUIDA"
    cnpj_consulta = digits(cliente.get("cnpj"))
    request_delay = max(0.0, float(getattr(settings, "request_delay_seconds", 0) or 0))
    pause_429_default = int(getattr(settings, "rate_limit_pause_seconds", 45) or 45)
    pause_429_max = int(getattr(settings, "rate_limit_max_pause_seconds", 300) or 300)
    max_429_retries = max(0, int(getattr(settings, "max_rate_limit_retries", 3) or 0))

    for nsu in range(nsu_start, nsu_start + max(0, int(max_consultas))):
        if cancel_checker and cancel_checker():
            result["mensagens"].append("Consulta cancelada pelo usuário.")
            status_final = "CANCELADA"
            break

        response = None
        rate_limit_hits = 0
        stop_current_cliente = False

        while response is None:
            if cancel_checker and cancel_checker():
                result["mensagens"].append("Consulta cancelada pelo usuário.")
                status_final = "CANCELADA"
                stop_current_cliente = True
                break

            if result["consultas_realizadas"] > 0 and request_delay > 0:
                if not sleep_with_cancel(request_delay, cancel_checker):
                    result["mensagens"].append("Consulta cancelada durante a pausa técnica entre requisições.")
                    status_final = "CANCELADA"
                    stop_current_cliente = True
                    break

            result["status"] = "EM CONSULTA"
            result["consultas_realizadas"] += 1
            result["ultimo_nsu_processado"] = nsu
            if progress_callback:
                progress_callback(result)

            try:
                candidate = client.get_dfe(nsu, cnpj_consulta=cnpj_consulta)
            except Exception as exc:  # noqa: BLE001
                result["mensagens"].append(f"NSU {nsu}: erro de conexão/consulta: {exc}")
                status_final = "ERRO"
                if progress_callback:
                    progress_callback(result)
                stop_current_cliente = True
                break

            if candidate.status_code == 429:
                rate_limit_hits += 1
                wait_seconds = retry_after_seconds(candidate.headers, pause_429_default, pause_429_max)
                if rate_limit_hits <= max_429_retries:
                    result["status"] = "AGUARDANDO LIMITE"
                    result["mensagens"].append(
                        f"NSU {nsu}: limite temporário de requisições do Portal Nacional (HTTP 429). "
                        f"Pausa técnica de {wait_seconds}s antes da tentativa {rate_limit_hits + 1}/{max_429_retries + 1}."
                    )
                    if progress_callback:
                        progress_callback(result)
                    if not sleep_with_cancel(wait_seconds, cancel_checker):
                        result["mensagens"].append("Consulta cancelada durante a pausa por limite de requisições.")
                        status_final = "CANCELADA"
                        stop_current_cliente = True
                        break
                    continue

                result["mensagens"].append(
                    f"NSU {nsu}: limite temporário de requisições do Portal Nacional (HTTP 429) persistiu após "
                    f"{rate_limit_hits} tentativa(s). A empresa foi interrompida para evitar novas recusas. "
                    "Aguarde alguns minutos e reinicie a busca; o sistema não avançou esse NSU como documento processado."
                )
                status_final = "ERRO"
                if progress_callback:
                    progress_callback(result)
                stop_current_cliente = True
                break

            response = candidate

        if stop_current_cliente:
            break

        if response is None:
            break

        if response.status_code in {204, 404}:
            empty_streak += 1
            if empty_streak == 1:
                result["mensagens"].append(f"NSU {nsu}: sem documento.")
            if progress_callback:
                progress_callback(result)
            if parar_apos_vazios and empty_streak >= int(parar_apos_vazios):
                result["mensagens"].append(f"Parado após {empty_streak} NSU(s) vazio(s).")
                if progress_callback:
                    progress_callback(result)
                break
            continue

        if response.status_code >= 400:
            result["mensagens"].append(http_error_message(nsu, response.status_code, response.text_preview))
            status_final = "ERRO"
            if progress_callback:
                progress_callback(result)
            break

        xmls = extract_xml_payloads(response)
        if not xmls:
            preview = clean_response_preview(response.text_preview, limit=260)
            result["mensagens"].append(f"NSU {nsu}: resposta recebida, mas nenhum XML foi extraído. Prévia: {preview or '-'}")
            status_final = "ERRO"
            if progress_callback:
                progress_callback(result)
            break

        empty_streak = 0
        result["documentos_recebidos"] += len(xmls)
        data = store.read()
        cliente = find_by_id(data.get("clientes", []), cliente_id) or cliente
        counts = process_xml_items(data, cliente, xmls, origem="ADN_API", nsu=nsu, inicio=inicio, fim=fim)
        result["importadas"] += int(counts.get("importadas", 0))
        result["atualizadas"] += int(counts.get("atualizadas", 0))
        result["ignoradas"] += int(counts.get("ignoradas", 0))
        result["fora_periodo"] += int(counts.get("fora_periodo", 0))
        result["mensagens"].extend(counts.get("mensagens", [])[:8])
        cliente["ultimo_nsu"] = max(int(cliente.get("ultimo_nsu") or 0), nsu)
        cliente["ultima_sincronizacao_em"] = now_iso()
        store.write(data)
        if progress_callback:
            progress_callback(result)

    data = store.read()
    cliente = find_by_id(data.get("clientes", []), cliente_id)
    if cliente:
        cliente["ultima_sincronizacao_em"] = now_iso()
    store.write(data)
    result["status"] = status_final
    if progress_callback:
        progress_callback(result)
    _finish_log(result)
    return result


def _finish_log(result: dict[str, Any]) -> None:
    result["finalizado_em"] = now_iso()
    data = store.read()
    for index, log in enumerate(data.get("logs", [])):
        if int(log.get("id", -1)) == int(result.get("id", -2)):
            data["logs"][index] = result
            break
    else:
        data.setdefault("logs", []).append(result)
    store.write(data)


def render_sync_results(results: list[dict[str, Any]], title: str) -> str:
    rows = []
    for r in results:
        msgs = "<br>".join(escape(str(m)) for m in r.get("mensagens", [])[-12:]) or "-"
        periodo = periodo_label(r.get("periodo_inicio"), r.get("periodo_fim"))
        rows.append(
            "<tr>"
            f"<td>{mask_cnpj(r.get('cnpj_cliente'))}</td>"
            f"<td>{escape(periodo)}</td>"
            f"<td>{pill(r.get('status',''), 'ok' if r.get('status') == 'CONCLUIDA' else 'err')}</td>"
            f"<td class='num'>{r.get('nsu_inicial')}</td>"
            f"<td class='num'>{r.get('ultimo_nsu_processado')}</td>"
            f"<td class='num'>{r.get('consultas_realizadas')}</td>"
            f"<td class='num'>{r.get('documentos_recebidos')}</td>"
            f"<td class='num'>{r.get('importadas')}</td>"
            f"<td class='num'>{r.get('atualizadas')}</td>"
            f"<td class='num'>{r.get('fora_periodo', 0)}</td>"
            f"<td class='num'>{r.get('ignoradas')}</td>"
            f"<td>{msgs}</td>"
            "</tr>"
        )
    body = "".join(rows) or "<tr><td colspan='12'>Nenhum cliente sincronizado.</td></tr>"
    html = f"""
      <section class="card">
        <h2>{escape(title)}</h2>
        <p class="muted small">Resultado da busca automática no ADN/NFS-e Nacional. A busca considera todos os municípios retornados para cada CNPJ.</p>
        <div class="actions" style="margin-top:14px"><a class="button ghost" href="/sincronizar">Nova busca</a><a class="button blue" href="/retencoes">Resumo por empresa</a><a class="button ghost" href="/relatorio">Notas detalhadas</a></div>
      </section>
      <div class="section-title"><h2>Resultado</h2></div>
      <div class="table-wrap"><table><thead><tr><th>CNPJ</th><th>Período</th><th>Status</th><th class="num">NSU inicial</th><th class="num">Último</th><th class="num">Consultas</th><th class="num">Docs</th><th class="num">Importadas</th><th class="num">Atualizadas</th><th class="num">Fora período</th><th class="num">Ignoradas</th><th>Mensagens</th></tr></thead><tbody>{body}</tbody></table></div>
    """
    return render_page(title, "sincronizar", html)


@app.get("/sincronizacao/ativa")
def sincronizacao_ativa() -> RedirectResponse:
    data = store.read()
    job = latest_active_job(data) or latest_job(data)
    if job:
        return RedirectResponse(f"/sincronizacao/progresso/{job['id']}", status_code=303)
    return RedirectResponse("/sincronizar", status_code=303)


@app.post("/ui/sincronizacao/{job_id}/parar")
def ui_parar_consulta(job_id: int) -> RedirectResponse:
    request_stop_job(job_id)
    return RedirectResponse(f"/sincronizacao/progresso/{job_id}", status_code=303)


@app.get("/api/sincronizacao/jobs/{job_id}")
def api_sync_job(job_id: int) -> dict[str, Any]:
    data = store.read()
    job = find_job(data, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Consulta não encontrada.")
    return json_safe(job)


@app.get("/sincronizacao/progresso/{job_id}", response_class=HTMLResponse)
def sync_progress_page(job_id: int) -> str:
    data = store.read()
    job = find_job(data, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Consulta não encontrada.")
    active = str(job.get("status") or "").upper() in ACTIVE_JOB_STATUSES
    progress = job_progress_percent(job)
    totals = job.get("totais", {})
    periodo = periodo_label(job.get("periodo_inicio"), job.get("periodo_fim"))
    current = job.get("empresa_atual") or {}

    rate_limit_seen = any(
        is_rate_limit_message(msg)
        for item in job.get("clientes", [])
        for msg in (item.get("mensagens", []) or [])
    )
    rate_limit_notice = ""
    if rate_limit_seen:
        rate_limit_notice = """
          <section class="notice rate-limit-panel">
            <strong>HTTP 429 identificado.</strong>
            <span>O Portal Nacional está limitando o ritmo de requisições. O sistema pausa, respeita Retry-After quando disponível e tenta novamente o mesmo NSU antes de marcar falha.</span>
          </section>
        """

    status_breakdown = {}
    for item in job.get("clientes", []):
        key = str(item.get("status") or "PENDENTE").upper()
        status_breakdown[key] = status_breakdown.get(key, 0) + 1
    status_chips = "".join(f"<span>{escape(k.title())}: <strong>{v}</strong></span>" for k, v in sorted(status_breakdown.items())) or "<span>Sem empresas na busca</span>"

    company_cards = []
    table_rows = []
    for item in job.get("clientes", []):
        msgs = [str(m) for m in (item.get("mensagens", []) or []) if str(m).strip()]
        display_msgs = msgs[-3:]
        msg_stack = "".join(f"<li>{escape(m)}</li>" for m in display_msgs) or "<li>Aguardando início da consulta.</li>"
        erro = item.get("erro") or ""
        if erro and not any(str(erro) in m for m in display_msgs):
            msg_stack += f"<li class='danger-text'>{escape(str(erro))}</li>"
        status_value = str(item.get("status") or "PENDENTE")
        company_cards.append(
            f"""
            <article class="progress-company-card">
              <div class="company-card-head">
                <div><strong>{escape(item.get('razao_social',''))}</strong><span>{mask_cnpj(item.get('cnpj'))}</span></div>
                {pill(status_value, sync_status_kind(status_value))}
              </div>
              <div class="company-card-metrics">
                <span><em>Último NSU</em><strong>{item.get('ultimo_nsu_processado', '')}</strong></span>
                <span><em>Docs</em><strong>{item.get('documentos_recebidos', 0)}</strong></span>
                <span><em>Novas</em><strong>{item.get('importadas', 0)}</strong></span>
                <span><em>Atualizadas</em><strong>{item.get('atualizadas', 0)}</strong></span>
              </div>
              <ul class="message-stack">{msg_stack}</ul>
            </article>
            """
        )
        table_rows.append(
            "<tr>"
            f"<td><strong>{escape(item.get('razao_social',''))}</strong><br><span class='small muted'>{mask_cnpj(item.get('cnpj'))}</span></td>"
            f"<td>{pill(status_value, sync_status_kind(status_value))}</td>"
            f"<td class='num'>{item.get('nsu_inicial', '')}</td>"
            f"<td class='num'>{item.get('ultimo_nsu_processado', '')}</td>"
            f"<td class='num'>{item.get('consultas_realizadas', 0)}</td>"
            f"<td class='num'>{item.get('documentos_recebidos', 0)}</td>"
            f"<td class='num'>{item.get('importadas', 0)}</td>"
            f"<td class='num'>{item.get('atualizadas', 0)}</td>"
            f"<td class='num'>{item.get('fora_periodo', 0)}</td>"
            f"<td>{'<br>'.join(escape(m) for m in display_msgs) or '-'}</td>"
            "</tr>"
        )

    card_grid = "".join(company_cards) or "<article class='progress-company-card'><strong>Nenhuma empresa na busca.</strong></article>"
    body_rows = "".join(table_rows) or "<tr><td colspan='10'>Nenhuma empresa na consulta.</td></tr>"

    stop_button = ""
    auto_refresh = ""
    if active:
        stop_button = f"""
          <form method="post" action="/ui/sincronizacao/{job_id}/parar" onsubmit="return confirm('Parar a consulta em andamento?');">
            <button class="stop strong-stop" type="submit">Parar busca</button>
          </form>
        """
        auto_refresh = "<script>setTimeout(() => window.location.reload(), 3000);</script>"

    body = f"""
      {rate_limit_notice}
      <section class="card progress-console">
        <div class="console-head">
          <div>
            <span class="small-eyebrow">Monitor de execução</span>
            <h2>{escape(str(job.get('mensagem_atual') or 'Acompanhamento da busca'))}</h2>
            <p class="muted small">Período: <strong>{escape(periodo)}</strong> · Empresas: <strong>{job.get('total_empresas', 0)}</strong> · Atualização automática enquanto ativo.</p>
          </div>
          <div class="operation-actions">
            {pill(job.get('status',''), sync_status_kind(job.get('status')))}
            {stop_button}
          </div>
        </div>
        <div class="progress-track"><span style="width:{progress}%"></span></div>
        <div class="progress-meta"><strong>{progress}%</strong><span>{job.get('empresas_processadas', 0)} de {job.get('total_empresas', 0)} empresas concluídas</span><em>Atual: {escape(current.get('razao_social') or '-')}</em></div>
        <div class="status-chip-row">{status_chips}</div>
      </section>

      <div class="kpi-strip">
        <div class="card metric blue"><span class="metric-label">Consultas</span><strong class="metric-value">{totals.get('consultas_realizadas', 0)}</strong><span class="metric-detail">requisições ao Portal Nacional</span></div>
        <div class="card metric accent"><span class="metric-label">Documentos</span><strong class="metric-value">{totals.get('documentos_recebidos', 0)}</strong><span class="metric-detail">XMLs/documentos retornados</span></div>
        <div class="card metric"><span class="metric-label">Notas novas</span><strong class="metric-value">{totals.get('importadas', 0)}</strong><span class="metric-detail">gravadas na base local</span></div>
        <div class="card metric dark"><span class="metric-label">Pendências</span><strong class="metric-value">{job.get('empresas_com_erro', 0)}</strong><span class="metric-detail">empresas com alerta</span></div>
      </div>

      <section class="card data-card">
        <div class="card-head">
          <div><span class="small-eyebrow">Empresas</span><h2>Status por CNPJ</h2></div>
          <div class="actions"><a class="button ghost compact-btn" href="/sincronizar">Nova busca</a><a class="button blue compact-btn" href="/retencoes">Ver retenções</a></div>
        </div>
        <div class="progress-company-grid">{card_grid}</div>
      </section>

      <details class="advanced-box">
        <summary>Detalhes técnicos de NSU e mensagens</summary>
        <div class="table-wrap"><table><thead><tr><th>Empresa</th><th>Status</th><th class="num">NSU inicial</th><th class="num">Último NSU</th><th class="num">Consultas</th><th class="num">Docs</th><th class="num">Importadas</th><th class="num">Atualizadas</th><th class="num">Fora período</th><th>Últimas mensagens</th></tr></thead><tbody>{body_rows}</tbody></table></div>
      </details>
      {auto_refresh}
    """
    return render_page("Execução da busca", "sincronizar", body, subtitle="Monitoramento por empresa, NSU e retorno do Portal Nacional.")

@app.post("/clientes/{cliente_id}/importar-xml")
async def importar_xml_cliente(cliente_id: int, files: list[UploadFile] = File(...)) -> dict[str, Any]:
    data = store.read()
    cliente = find_by_id(data.get("clientes", []), cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    total = {"notas_lidas": 0, "importadas": 0, "atualizadas": 0, "ignoradas": 0, "mensagens": []}
    for upload in files:
        content = await upload.read()
        xmls = expand_upload(upload.filename or "upload.xml", content)
        counts = process_xml_items(data, cliente, xmls, origem="UPLOAD_XML")
        for key in ["notas_lidas", "importadas", "atualizadas", "ignoradas"]:
            total[key] += int(counts.get(key, 0))
        total["mensagens"].extend(counts.get("mensagens", []))
    store.write(data)
    return total


# ---------------------------------------------------------------------------
# Relatórios
# ---------------------------------------------------------------------------



RESUMO_EMPRESAS_COLUMNS = [
    ("razao_social", "Empresa", None),
    ("cnpj", "CNPJ", None),
    ("municipios_notas", "Municípios nas notas", None),
    ("prestado.qtd", "Prestado - autorizadas", None),
    ("prestado.qtd_canceladas", "Prestado - canceladas", None),
    ("prestado.qtd_substituidas", "Prestado - substituídas", None),
    ("prestado.valor_servico", "Prestado - serviços", "money"),
    ("prestado.base_calculo", "Prestado - base ISS", "money"),
    ("prestado.valor_iss", "Prestado - ISS", "money"),
    ("prestado.valor_iss_retido", "Prestado - ISS retido", "money"),
    ("prestado.valor_pis", "Prestado - PIS retido", "money"),
    ("prestado.valor_cofins", "Prestado - COFINS retido", "money"),
    ("prestado.valor_inss", "Prestado - INSS/CP retido", "money"),
    ("prestado.valor_ir", "Prestado - IRRF retido", "money"),
    ("prestado.valor_csll", "Prestado - CSLL retida", "money"),
    ("prestado.retencao_pis_cofins_csll_base", "Prestado - vRetCSLL agregado", "money"),
    ("prestado.outras_retencoes", "Prestado - outras", "money"),
    ("prestado.total_retencoes_federais", "Prestado - ret. federais", "money"),
    ("prestado.total_retencoes", "Prestado - total retido", "money"),
    ("prestado.valor_liquido", "Prestado - líquido", "money"),
    ("tomado.qtd", "Tomado - autorizadas", None),
    ("tomado.qtd_canceladas", "Tomado - canceladas", None),
    ("tomado.qtd_substituidas", "Tomado - substituídas", None),
    ("tomado.valor_servico", "Tomado - serviços", "money"),
    ("tomado.base_calculo", "Tomado - base ISS", "money"),
    ("tomado.valor_iss", "Tomado - ISS", "money"),
    ("tomado.valor_iss_retido", "Tomado - ISS retido", "money"),
    ("tomado.valor_pis", "Tomado - PIS retido", "money"),
    ("tomado.valor_cofins", "Tomado - COFINS retido", "money"),
    ("tomado.valor_inss", "Tomado - INSS/CP retido", "money"),
    ("tomado.valor_ir", "Tomado - IRRF retido", "money"),
    ("tomado.valor_csll", "Tomado - CSLL retida", "money"),
    ("tomado.retencao_pis_cofins_csll_base", "Tomado - vRetCSLL agregado", "money"),
    ("tomado.outras_retencoes", "Tomado - outras", "money"),
    ("tomado.total_retencoes_federais", "Tomado - ret. federais", "money"),
    ("tomado.total_retencoes", "Tomado - total retido", "money"),
    ("tomado.valor_liquido", "Tomado - líquido", "money"),
    ("qtd_canceladas", "Canceladas", None),
    ("qtd_substituidas", "Substituídas", None),
]


def nested_get(row: dict[str, Any], path: str) -> Any:
    current: Any = row
    for part in path.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


ROLE_SUMMARY_FIELDS = [
    ("qtd", "Notas autorizadas", "int"),
    ("qtd_canceladas", "Canceladas", "int"),
    ("qtd_substituidas", "Substituídas", "int"),
    ("valor_servico", "Valor dos serviços", "money"),
    ("valor_deducoes", "Deduções", "money"),
    ("total_descontos", "Descontos", "money"),
    ("base_calculo", "Base ISS", "money"),
    ("valor_iss", "ISS destacado", "money"),
    ("valor_iss_retido", "ISS retido", "money"),
    ("valor_pis", "PIS retido", "money"),
    ("valor_cofins", "COFINS retido", "money"),
    ("valor_csll", "CSLL retida", "money"),
    ("valor_ir", "IRRF retido", "money"),
    ("valor_inss", "INSS/CP retido", "money"),
    ("retencao_pis_cofins_csll_base", "vRetCSLL agregado", "money"),
    ("outras_retencoes", "Outras retenções", "money"),
    ("total_retencoes_federais", "Retenções federais", "money"),
    ("total_retencoes", "Total retido", "money"),
    ("valor_liquido", "Valor líquido", "money"),
]


def render_role_summary_table(rows: list[dict[str, Any]], totals: dict[str, Any], role: str, title: str, note: str) -> str:
    role = role.lower()
    head = ["Empresa", "CNPJ", "Municípios"] + [label for _key, label, _kind in ROLE_SUMMARY_FIELDS]
    header = "".join(f"<th class='num'>{escape(label)}</th>" if label not in {"Empresa", "CNPJ", "Municípios"} else f"<th>{escape(label)}</th>" for label in head)
    body_rows: list[str] = []
    for row in rows:
        role_data = row.get(role, {})
        cells = [
            f"<td><strong>{escape(str(row.get('razao_social') or ''))}</strong></td>",
            f"<td class='nowrap'>{mask_cnpj(str(row.get('cnpj') or ''))}</td>",
            f"<td>{escape(str(row.get('municipios_notas') or '-'))}</td>",
        ]
        for key, _label, kind in ROLE_SUMMARY_FIELDS:
            value = role_data.get(key)
            if kind == "money":
                cells.append(f"<td class='num'>{fmt_money(value)}</td>")
            else:
                cells.append(f"<td class='num'>{escape(str(value or 0))}</td>")
        body_rows.append("<tr>" + "".join(cells) + "</tr>")
    if not body_rows:
        body_rows.append(f"<tr><td colspan='{len(head)}'>Nenhuma empresa encontrada para o filtro.</td></tr>")
    footer_cells = ["<td>TOTAL GERAL</td>", "<td></td>", "<td></td>"]
    for key, _label, kind in ROLE_SUMMARY_FIELDS:
        value = totals.get(role, {}).get(key)
        footer_cells.append(f"<td class='num'>{fmt_money(value) if kind == 'money' else escape(str(value or 0))}</td>")
    footer = "<tr>" + "".join(footer_cells) + "</tr>"
    return f"""
      <section class="card role-summary-card">
        <div class="role-summary-head">
          <div><span class="eyebrow small-eyebrow">{escape(title)}</span><h2>{escape(title)}</h2><p class="muted small">{escape(note)}</p></div>
          <div class="role-kpis"><span><strong>{totals.get(role, {}).get('qtd', 0)}</strong> autorizadas</span><span><strong>{fmt_money(totals.get(role, {}).get('valor_servico'))}</strong> serviços</span><span><strong>{fmt_money(totals.get(role, {}).get('total_retencoes'))}</strong> retido</span></div>
        </div>
        <div class="table-wrap role-table"><table><thead><tr>{header}</tr></thead><tbody>{''.join(body_rows)}</tbody><tfoot>{footer}</tfoot></table></div>
      </section>
    """


def render_resumo_empresas_table(rows: list[dict[str, Any]], totals: dict[str, Any]) -> str:
    header = "".join(f"<th class='num'>{escape(label)}</th>" if kind == "money" or label.endswith("qtd") or label in {"Canceladas", "Substituídas"} else f"<th>{escape(label)}</th>" for path, label, kind in RESUMO_EMPRESAS_COLUMNS)
    body_rows: list[str] = []
    for row in rows:
        cells: list[str] = []
        for path, label, kind in RESUMO_EMPRESAS_COLUMNS:
            value = nested_get(row, path)
            if path == "cnpj":
                cells.append(f"<td class='nowrap'>{mask_cnpj(str(value or ''))}</td>")
            elif kind == "money":
                cells.append(f"<td class='num'>{fmt_money(value)}</td>")
            elif path.endswith("qtd") or label in {"Canceladas", "Substituídas"}:
                cells.append(f"<td class='num'>{escape(str(value or 0))}</td>")
            else:
                cells.append(f"<td>{escape(str(value or ''))}</td>")
        body_rows.append("<tr>" + "".join(cells) + "</tr>")
    if not body_rows:
        body_rows.append(f"<tr><td colspan='{len(RESUMO_EMPRESAS_COLUMNS)}'>Nenhuma empresa cadastrada.</td></tr>")

    total_row = {
        "razao_social": "TOTAL GERAL",
        "cnpj": "",
        "municipios_notas": "",
        "qtd_canceladas": totals.get("qtd_canceladas", 0),
        "qtd_substituidas": totals.get("qtd_substituidas", 0),
        "prestado": totals.get("prestado", {}),
        "tomado": totals.get("tomado", {}),
    }
    footer_cells: list[str] = []
    for path, label, kind in RESUMO_EMPRESAS_COLUMNS:
        value = nested_get(total_row, path)
        if path == "razao_social":
            footer_cells.append("<td>TOTAL GERAL</td>")
        elif path in {"cnpj", "municipios_notas"}:
            footer_cells.append("<td></td>")
        elif kind == "money":
            footer_cells.append(f"<td class='num'>{fmt_money(value)}</td>")
        else:
            footer_cells.append(f"<td class='num'>{escape(str(value or 0))}</td>")
    footer = "<tr>" + "".join(footer_cells) + "</tr>"
    return f"<div class='table-wrap summary-table'><table><thead><tr>{header}</tr></thead><tbody>{''.join(body_rows)}</tbody><tfoot>{footer}</tfoot></table></div>"


@app.get("/resumo-empresas", response_class=HTMLResponse)
@app.get("/retencoes", response_class=HTMLResponse)
def resumo_empresas(
    inicio: str | None = Query(default=None),
    fim: str | None = Query(default=None),
    status: str = Query(default="TODOS"),
    data_base: str = Query(default="competencia"),
    cliente_ids: list[int] = Query(default=[]),
    reprocessadas: int | None = Query(default=None),
    sem_xml: int | None = Query(default=None),
    falhas: int | None = Query(default=None),
) -> str:
    data = store.read()
    selected = normalize_id_list(cliente_ids)
    base = normalize_date_base(data_base)
    inicio_d = parse_date_filter(inicio)
    fim_d = parse_date_filter(fim)
    rows = resumo_empresas_rows(data, inicio=inicio_d, fim=fim_d, status=status, cliente_ids=selected, data_base=base)
    totals = resumo_empresas_totais(rows)

    status_options = html_select_options([
        ("TODOS", "Todos os status"),
        ("AUTORIZADA", "Somente autorizadas"),
        ("CANCELADA", "Somente canceladas"),
        ("SUBSTITUIDA", "Somente substituídas"),
    ], status)
    query = build_query(inicio=inicio, fim=fim, status=status, data_base=base, cliente_ids=selected)
    selecionadas = f"{len(selected)} selecionada(s)" if selected else "todas as empresas"
    periodo = periodo_label(inicio, fim, base)
    data_base_select = date_base_options_html(base)

    reprocess_notice = ""
    if reprocessadas is not None:
        reprocess_notice = f"""
          <section class="notice success">
            <strong>Retenções recalculadas.</strong>
            <span>{int(reprocessadas or 0)} nota(s) atualizada(s), {int(sem_xml or 0)} sem XML local e {int(falhas or 0)} com alerta.</span>
          </section>
        """

    filter_form = f"""
      <section class="toolbar-card filter-panel">
        <div class="toolbar-copy">
          <span class="small-eyebrow">Filtros do relatório</span>
          <h2>Mapa de retenções</h2>
          <p>Defina período, data-base, status e empresas para visualizar valores retidos por tributo.</p>
        </div>
        <form class="filter-card" method="get" action="/retencoes">
          <div class="form-row report-filter-row">
            <div><label>Início</label><input type="date" name="inicio" value="{escape(inicio or '')}"></div>
            <div><label>Fim</label><input type="date" name="fim" value="{escape(fim or '')}"></div>
            <div><label>Data-base</label><select name="data_base">{data_base_select}</select></div>
            <div><label>Status</label><select name="status">{status_options}</select></div>
          </div>
          <details class="filter-details">
            <summary>Empresas do relatório</summary>
            <p class="muted small">Sem marcação, o sistema considera todas as empresas ativas e notas já importadas.</p>
            {cliente_checkboxes(data.get('clientes', []), selected)}
          </details>
          <div class="command-actions">
            <button type="submit">Aplicar filtros</button>
            <a class="button ghost" href="/retencoes">Limpar</a>
            <a class="button blue" href="/relatorios/resumo-empresas.xlsx?{query}">Exportar Excel</a>
            <a class="button ghost" href="/relatorios/resumo-empresas.csv?{query}">CSV</a>
          </div>
        </form>
      </section>
    """

    cards = f"""
      <div class="kpi-strip">
        <div class="card metric blue"><span class="metric-label">Prestado autorizado</span><strong class="metric-value">{fmt_money(totals['prestado']['valor_servico'])}</strong><span class="metric-detail">{totals['prestado']['qtd']} nota(s) emitida(s)</span></div>
        <div class="card metric"><span class="metric-label">Tomado autorizado</span><strong class="metric-value">{fmt_money(totals['tomado']['valor_servico'])}</strong><span class="metric-detail">{totals['tomado']['qtd']} nota(s) recebida(s)</span></div>
        <div class="card metric accent"><span class="metric-label">Total retido</span><strong class="metric-value">{fmt_money(totals['prestado']['total_retencoes'] + totals['tomado']['total_retencoes'])}</strong><span class="metric-detail">ISS + federais + outras</span></div>
        <div class="card metric dark"><span class="metric-label">PIS/COFINS/CSLL</span><strong class="metric-value">{fmt_money(totals['prestado']['valor_pis'] + totals['prestado']['valor_cofins'] + totals['prestado']['valor_csll'] + totals['tomado']['valor_pis'] + totals['tomado']['valor_cofins'] + totals['tomado']['valor_csll'])}</strong><span class="metric-detail">segregado por retenção</span></div>
      </div>
    """

    def retention_rows_html(role_total: dict[str, Any]) -> str:
        rows_ret = [
            ("ISS retido", "valor_iss_retido"),
            ("PIS retido", "valor_pis"),
            ("COFINS retido", "valor_cofins"),
            ("CSLL retida", "valor_csll"),
            ("IRRF retido", "valor_ir"),
            ("INSS/CP retido", "valor_inss"),
            ("Outras retenções", "outras_retencoes"),
            ("Total retido", "total_retencoes"),
        ]
        return "".join(
            f"<tr><td>{escape(label)}</td><td class='num'>{fmt_money(role_total.get(key))}</td></tr>"
            for label, key in rows_ret
        )

    retention_breakdown = f"""
      <section class="card data-card retention-matrix">
        <div class="card-head">
          <div>
            <span class="small-eyebrow">Livro de retenções</span>
            <h2>Retido por tributo e natureza</h2>
          </div>
          <form method="post" action="/ui/reprocessar-retencoes" onsubmit="return confirm('Recalcular as retenções das notas já salvas usando os XMLs locais?');">
            <button class="ghost compact-btn" type="submit">Recalcular XMLs</button>
          </form>
        </div>
        <div class="grid two">
          <section class="ledger-card">
            <h3>Notas prestadas</h3>
            <p class="muted small">Retenções sofridas nas notas emitidas pela empresa.</p>
            <div class="table-wrap compact"><table><thead><tr><th>Retenção</th><th class="num">Valor</th></tr></thead><tbody>{retention_rows_html(totals['prestado'])}</tbody></table></div>
          </section>
          <section class="ledger-card">
            <h3>Notas tomadas</h3>
            <p class="muted small">Retenções informadas nos serviços contratados/recebidos.</p>
            <div class="table-wrap compact"><table><thead><tr><th>Retenção</th><th class="num">Valor</th></tr></thead><tbody>{retention_rows_html(totals['tomado'])}</tbody></table></div>
          </section>
        </div>
        <div class="retention-disclaimer">
          <strong>Critério técnico</strong>
          <span>vPIS e vCOFINS permanecem como apuração própria. Retenções sociais usam tags explícitas ou tpRetPisCofins + vRetCSLL, com distribuição por PIS, COFINS e CSLL conforme o tipo indicado no XML.</span>
        </div>
      </section>
    """

    prestado_table = render_role_summary_table(
        rows,
        totals,
        "prestado",
        "Prestadas por empresa",
        "Faturamento, retenções sofridas e valor líquido das notas emitidas.",
    )
    tomado_table = render_role_summary_table(
        rows,
        totals,
        "tomado",
        "Tomadas por empresa",
        "Serviços recebidos, retenções informadas e documentos tomados.",
    )

    body = f"""
      {filter_form}
      {reprocess_notice}
      <div class="context-line">
        <span>Escopo: <strong>{escape(selecionadas)}</strong></span>
        <span>Período: <strong>{escape(periodo)}</strong></span>
        <span>Data-base: <strong>{escape(date_base_label(base))}</strong></span>
        <span>Status: <strong>{escape(status)}</strong></span>
      </div>
      {cards}
      {retention_breakdown}
      <section class="card data-card">
        <div class="card-head">
          <div><span class="small-eyebrow">Resumo por empresa</span><h2>Prestadas e tomadas</h2></div>
          <span class="muted small">Valores monetários consideram notas autorizadas; canceladas e substituídas ficam separadas.</span>
        </div>
        <div class="summary-stack">
          {prestado_table}
          {tomado_table}
        </div>
      </section>
    """
    return render_page("Retenções", "retencoes", body, subtitle="Conferência por tributo, empresa, natureza e data-base.")

@app.get("/relatorios/resumo-empresas.csv")
def relatorio_resumo_empresas_csv(
    inicio: str | None = None,
    fim: str | None = None,
    status: str = "TODOS",
    data_base: str = "competencia",
    cliente_ids: list[int] = Query(default=[]),
) -> StreamingResponse:
    data = store.read()
    base = normalize_date_base(data_base)
    rows = resumo_empresas_rows(data, inicio=parse_date_filter(inicio), fim=parse_date_filter(fim), status=status, cliente_ids=normalize_id_list(cliente_ids), data_base=base)
    totals = resumo_empresas_totais(rows)
    buffer = StringIO()
    buffer.write(";".join(label for _path, label, _kind in RESUMO_EMPRESAS_COLUMNS) + "\n")
    for row in rows:
        values: list[str] = []
        for path, _label, kind in RESUMO_EMPRESAS_COLUMNS:
            value = nested_get(row, path)
            if path == "cnpj":
                values.append(mask_cnpj(str(value or "")))
            elif kind == "money":
                values.append(fmt_num(value))
            else:
                values.append(str(value or "").replace(";", ",").replace("\n", " "))
        buffer.write(";".join(values) + "\n")
    total_row = {"razao_social": "TOTAL GERAL", "cnpj": "", "municipios_notas": "", "qtd_canceladas": totals["qtd_canceladas"], "qtd_substituidas": totals["qtd_substituidas"], "prestado": totals["prestado"], "tomado": totals["tomado"]}
    values = []
    for path, _label, kind in RESUMO_EMPRESAS_COLUMNS:
        value = nested_get(total_row, path)
        values.append(fmt_num(value) if kind == "money" else str(value or ""))
    buffer.write(";".join(values) + "\n")
    return StreamingResponse(
        iter([buffer.getvalue().encode("utf-8-sig")]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=ora_nfse_resumo_empresas.csv"},
    )


@app.get("/relatorios/resumo-empresas.xlsx")
def relatorio_resumo_empresas_xlsx(
    inicio: str | None = None,
    fim: str | None = None,
    status: str = "TODOS",
    data_base: str = "competencia",
    cliente_ids: list[int] = Query(default=[]),
) -> StreamingResponse:
    data = store.read()
    selected = normalize_id_list(cliente_ids)
    base = normalize_date_base(data_base)
    rows = resumo_empresas_rows(data, inicio=parse_date_filter(inicio), fim=parse_date_filter(fim), status=status, cliente_ids=selected, data_base=base)
    totals = resumo_empresas_totais(rows)
    wb = Workbook()
    wb.remove(wb.active)

    def append_role_sheet(sheet_name: str, role: str) -> None:
        ws = wb.create_sheet(sheet_name)
        headers = ["Empresa", "CNPJ", "Municípios"] + [label for _key, label, _kind in ROLE_SUMMARY_FIELDS]
        ws.append(headers)
        for row in rows:
            role_data = row.get(role, {})
            excel_row: list[Any] = [row.get("razao_social"), mask_cnpj(str(row.get("cnpj") or "")), row.get("municipios_notas")]
            for key, _label, kind in ROLE_SUMMARY_FIELDS:
                value = role_data.get(key)
                excel_row.append(float(dec(value)) if kind == "money" else int(value or 0))
            ws.append(excel_row)
        total_row: list[Any] = ["TOTAL GERAL", "", ""]
        for key, _label, kind in ROLE_SUMMARY_FIELDS:
            value = totals.get(role, {}).get(key)
            total_row.append(float(dec(value)) if kind == "money" else int(value or 0))
        ws.append(total_row)

    append_role_sheet("Prestados", "prestado")
    append_role_sheet("Tomados", "tomado")

    ws_tot = wb.create_sheet("Totalizadores")
    ws_tot.append(["ORA NFS-e Automático", "Resumo por empresa"])
    ws_tot.append(["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws_tot.append(["Período", periodo_label(inicio, fim, base)])
    ws_tot.append(["Data-base", date_base_label(base)])
    ws_tot.append(["Status", status])
    ws_tot.append(["Empresas selecionadas", "Todas" if not selected else len(selected)])
    ws_tot.append([])
    ws_tot.append(["Campo", "Serviços prestados", "Serviços tomados"])
    labels = [
        ("Notas autorizadas", "qtd"),
        ("Canceladas", "qtd_canceladas"),
        ("Substituídas", "qtd_substituidas"),
        ("Valor dos serviços", "valor_servico"),
        ("Deduções", "valor_deducoes"),
        ("Descontos", "total_descontos"),
        ("Base ISS", "base_calculo"),
        ("ISS destacado", "valor_iss"),
        ("ISS retido", "valor_iss_retido"),
        ("PIS retido", "valor_pis"),
        ("COFINS retido", "valor_cofins"),
        ("INSS/CP retido", "valor_inss"),
        ("IRRF retido", "valor_ir"),
        ("CSLL retida", "valor_csll"),
        ("PIS/COFINS/CSLL agregado no XML", "retencao_pis_cofins_csll_base"),
        ("PIS apurado XML — não retido", "valor_pis_apurado"),
        ("COFINS apurado XML — não retido", "valor_cofins_apurado"),
        ("Outras retenções", "outras_retencoes"),
        ("Retenções federais", "total_retencoes_federais"),
        ("Total retido", "total_retencoes"),
        ("Valor líquido", "valor_liquido"),
    ]
    for label, key in labels:
        pval = totals["prestado"].get(key)
        tval = totals["tomado"].get(key)
        if key.startswith("qtd"):
            ws_tot.append([label, int(pval or 0), int(tval or 0)])
        else:
            ws_tot.append([label, float(dec(pval)), float(dec(tval))])

    ws_conf = wb.create_sheet("Conferência")
    ws_conf.append(["Empresa", "CNPJ", "Prestado - serviços", "Prestado - retido", "Prestado - líquido", "Tomado - serviços", "Tomado - retido", "Tomado - líquido", "Municípios"])
    for row in rows:
        ws_conf.append([
            row.get("razao_social"),
            mask_cnpj(str(row.get("cnpj") or "")),
            float(dec(row.get("prestado", {}).get("valor_servico"))),
            float(dec(row.get("prestado", {}).get("total_retencoes"))),
            float(dec(row.get("prestado", {}).get("valor_liquido"))),
            float(dec(row.get("tomado", {}).get("valor_servico"))),
            float(dec(row.get("tomado", {}).get("total_retencoes"))),
            float(dec(row.get("tomado", {}).get("valor_liquido"))),
            row.get("municipios_notas"),
        ])

    # Estilo executivo ORA nas abas exportadas.
    for sheet in wb.worksheets:
        if sheet.max_row:
            header_row = 1
            if sheet.title == "Totalizadores":
                header_row = 8
                sheet["A1"].font = Font(bold=True, color="061F3F", size=16)
            for cell in sheet[header_row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="061F3F")
        for col in range(1, min(sheet.max_column, 40) + 1):
            letter = get_column_letter(col)
            sheet.column_dimensions[letter].width = min(max(12, max((len(str(sheet.cell(row=r, column=col).value or "")) for r in range(1, min(sheet.max_row, 80) + 1)), default=12) + 2), 46)
            for row_idx in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=col)
                if isinstance(cell.value, float):
                    cell.number_format = 'R$ #,##0.00'
        sheet.freeze_panes = "A2" if sheet.title != "Totalizadores" else "A9"
        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ora_nfse_resumo_empresas_separado.xlsx"},
    )


@app.get("/relatorio", response_class=HTMLResponse)
def relatorio(
    cliente_id: int | None = Query(default=None),
    inicio: str | None = Query(default=None),
    fim: str | None = Query(default=None),
    papel: str = Query(default="PRESTADOR"),
    status: str = Query(default="TODOS"),
    data_base: str = Query(default="competencia"),
) -> str:
    data = store.read()
    base = normalize_date_base(data_base)
    inicio_d = parse_date_filter(inicio)
    fim_d = parse_date_filter(fim)
    notes = filtered_notes(data, cliente_id=cliente_id, inicio=inicio_d, fim=fim_d, papel=papel, status=status, data_base=base)
    totals = totals_for_notes(notes)
    cliente = find_by_id(data.get("clientes", []), cliente_id)

    cliente_opts = [("", "Todos os clientes")] + [(c["id"], f"{c.get('razao_social','')} · {mask_cnpj(c.get('cnpj'))}") for c in data.get("clientes", [])]
    papel_opts = [("PRESTADOR", "Notas emitidas / faturamento"), ("TOMADOR", "Serviços tomados"), ("TODOS", "Emitidas e tomadas")]
    status_opts = [("TODOS", "Todos os status"), ("AUTORIZADA", "Autorizadas"), ("CANCELADA", "Canceladas"), ("SUBSTITUIDA", "Substituídas")]
    data_base_select = date_base_options_html(base)
    query = build_query(cliente_id=cliente_id, inicio=inicio, fim=fim, papel=papel, status=status, data_base=base)

    filter_form = f"""
      <section class="toolbar-card filter-panel">
        <div class="toolbar-copy">
          <span class="small-eyebrow">Filtros de auditoria</span>
          <h2>Notas fiscais detalhadas</h2>
          <p>Use a data-base de competência ou emissão para revisar XMLs, status, valores e retenções.</p>
        </div>
        <form class="filter-card" method="get" action="/relatorio">
          <div class="form-row report-filter-row">
            <div><label>Cliente</label><select name="cliente_id">{html_select_options(cliente_opts, cliente_id or '')}</select></div>
            <div><label>Início</label><input type="date" name="inicio" value="{escape(inicio or '')}"></div>
            <div><label>Fim</label><input type="date" name="fim" value="{escape(fim or '')}"></div>
            <div><label>Data-base</label><select name="data_base">{data_base_select}</select></div>
            <div><label>Natureza</label><select name="papel">{html_select_options(papel_opts, papel)}</select></div>
            <div><label>Status</label><select name="status">{html_select_options(status_opts, status)}</select></div>
          </div>
          <div class="command-actions">
            <button type="submit">Aplicar filtros</button>
            <a class="button ghost" href="/relatorio">Limpar</a>
            <a class="button blue" href="/relatorios/notas.xlsx?{query}">Exportar Excel</a>
          </div>
        </form>
      </section>
    """

    metric_cards = f"""
      <div class="kpi-strip">
        <div class="card metric blue"><span class="metric-label">Serviços exibidos</span><strong class="metric-value">{fmt_money(totals['valor_servico'])}</strong><span class="metric-detail">notas autorizadas no filtro</span></div>
        <div class="card metric accent"><span class="metric-label">Total retido</span><strong class="metric-value">{fmt_money(totals['total_retencoes'])}</strong><span class="metric-detail">ISS + federais + outras</span></div>
        <div class="card metric"><span class="metric-label">Líquido</span><strong class="metric-value">{fmt_money(totals['valor_liquido'])}</strong><span class="metric-detail">valor líquido autorizado</span></div>
        <div class="card metric dark"><span class="metric-label">Notas no filtro</span><strong class="metric-value">{totals['qtd_total']}</strong><span class="metric-detail">{totals['qtd_autorizadas']} autorizadas · {totals['qtd_canceladas']} canceladas</span></div>
      </div>
    """

    confer_rows = [
        ("Valor dos serviços", totals["valor_servico"]),
        ("Deduções", totals["valor_deducoes"]),
        ("Descontos", totals["total_descontos"]),
        ("Base de cálculo ISS", totals["base_calculo"]),
        ("ISS destacado", totals["valor_iss"]),
        ("ISS retido", totals["valor_iss_retido"]),
        ("PIS retido", totals["valor_pis"]),
        ("COFINS retido", totals["valor_cofins"]),
        ("CSLL retida", totals["valor_csll"]),
        ("IRRF retido", totals["valor_ir"]),
        ("INSS/CP retido", totals["valor_inss"]),
        ("PIS/COFINS/CSLL agregado no XML", totals.get("retencao_pis_cofins_csll_base", Decimal("0"))),
        ("PIS apurado no XML — não tratado como retenção", totals.get("valor_pis_apurado", Decimal("0"))),
        ("COFINS apurado no XML — não tratado como retenção", totals.get("valor_cofins_apurado", Decimal("0"))),
        ("Outras retenções", totals["outras_retencoes"]),
        ("Retenções federais", totals["total_retencoes_federais"]),
        ("Total retido", totals["total_retencoes"]),
        ("Valor líquido", totals["valor_liquido"]),
    ]
    confer_html = "".join(f"<tr><td>{escape(label)}</td><td class='num'>{fmt_money(value)}</td></tr>" for label, value in confer_rows)

    month_html = "".join(
        f"<tr><td>{escape(row['mes'])}</td><td class='num'>{row['qtd_autorizadas']}</td><td class='num'>{row['qtd_canceladas']}</td><td class='num'>{fmt_money(row['faturamento_bruto'])}</td><td class='num'>{fmt_money(row['retencoes_faturamento'])}</td><td class='num'>{fmt_money(row['faturamento_liquido'])}</td></tr>"
        for row in monthly_rows(notes, base)
    ) or "<tr><td colspan='6'>Sem notas no filtro.</td></tr>"

    detail_html = render_detail_table(notes, totals)
    body = f"""
      {filter_form}
      <div class="context-line">
        <span>Cliente: <strong>{escape(cliente_label(cliente))}</strong></span>
        <span>Período: <strong>{escape(periodo_label(inicio, fim, base))}</strong></span>
        <span>Data-base: <strong>{escape(date_base_label(base))}</strong></span>
        <span>Natureza: <strong>{escape(papel)}</strong></span>
      </div>
      {metric_cards}

      <section class="notice system-note">
        <strong>Critério fiscal</strong>
        <span>vPIS/vCOFINS ficam como apuração própria quando o XML traz essa estrutura. Retenção na fonte aparece nos campos PIS retido, COFINS retido e CSLL retida.</span>
      </section>

      <section class="grid two">
        <section class="card data-card"><div class="card-head"><div><span class="small-eyebrow">Totalizadores</span><h2>Conferência do filtro</h2></div></div><div class="table-wrap compact"><table><thead><tr><th>Campo</th><th class="num">Total</th></tr></thead><tbody>{confer_html}</tbody></table></div></section>
        <section class="card data-card"><div class="card-head"><div><span class="small-eyebrow">Resumo mensal</span><h2>Movimento por mês</h2></div></div><div class="table-wrap compact"><table><thead><tr><th>Mês</th><th class="num">Autorizadas</th><th class="num">Canceladas</th><th class="num">Serviços</th><th class="num">Retenções</th><th class="num">Líquido</th></tr></thead><tbody>{month_html}</tbody></table></div></section>
      </section>

      <section class="card data-card">
        <div class="card-head">
          <div><span class="small-eyebrow">Tabela analítica</span><h2>Notas fiscais</h2></div>
          <div class="actions"><a class="button ghost compact-btn" href="/relatorios/notas.csv?{query}">CSV</a><a class="button blue compact-btn" href="/relatorios/notas.xlsx?{query}">Excel</a></div>
        </div>
        {detail_html}
      </section>
    """
    return render_page("Notas", "relatorio", body, subtitle="Auditoria detalhada de XMLs, valores e retenções.")

def render_detail_table(notes: list[dict[str, Any]], totals: dict[str, Any]) -> str:
    header = "".join(f"<th>{escape(label)}</th>" if key not in MONEY_FIELDS else f"<th class='num'>{escape(label)}</th>" for key, label in DETAIL_COLUMNS)
    rows = []
    for note in notes:
        cells = []
        for key, _label in DETAIL_COLUMNS:
            value = note.get(key)
            if key in {"cnpj_prestador", "cnpj_tomador"}:
                cells.append(f"<td class='nowrap'>{mask_cnpj(str(value or ''))}</td>")
            elif key in {"data_emissao", "competencia"}:
                cells.append(f"<td class='nowrap'>{fmt_date(str(value or ''))}</td>")
            elif key == "status":
                cells.append(f"<td>{pill(str(value or ''))}</td>")
            elif key == "iss_retido":
                cells.append(f"<td>{'Sim' if value is True else 'Não' if value is False else '-'}</td>")
            elif key == "iss_retido_tipo":
                cells.append(f"<td>{escape(iss_retido_tipo_label(value))}</td>")
            elif key == "retencao_pis_cofins_csll_tipo":
                cells.append(f"<td>{escape(social_retencao_tipo_label(value))}</td>")
            elif key in MONEY_FIELDS:
                cells.append(f"<td class='num'>{fmt_money(value)}</td>")
            elif key == "aliquota_iss":
                cells.append(f"<td class='num'>{escape(str(value or ''))}</td>")
            elif key == "chave_acesso":
                cells.append(f"<td class='small'>{escape(str(value or ''))}</td>")
            else:
                text = str(value or "")
                if key == "discriminacao" and len(text) > 180:
                    text = text[:177] + "..."
                cells.append(f"<td>{escape(text)}</td>")
        rows.append("<tr>" + "".join(cells) + "</tr>")
    if not rows:
        rows.append(f"<tr><td colspan='{len(DETAIL_COLUMNS)}'>Nenhuma nota encontrada para o filtro.</td></tr>")
    footer_cells = []
    for key, _label in DETAIL_COLUMNS:
        if key == "discriminacao":
            footer_cells.append("<td>Total das autorizadas exibidas</td>")
        elif key in totals:
            footer_cells.append(f"<td class='num'>{fmt_money(totals[key])}</td>")
        elif key == "valor_servico":
            footer_cells.append(f"<td class='num'>{fmt_money(totals['valor_servico'])}</td>")
        else:
            footer_cells.append("<td></td>")
    footer = "<tr>" + "".join(footer_cells) + "</tr>"
    return f"<div class='table-wrap'><table><thead><tr>{header}</tr></thead><tbody>{''.join(rows)}</tbody><tfoot>{footer}</tfoot></table></div>"


def _notes_for_export(cliente_id: int | None, inicio: str | None, fim: str | None, papel: str, status: str, data_base: str = "competencia") -> list[dict[str, Any]]:
    data = store.read()
    base = normalize_date_base(data_base)
    return filtered_notes(data, cliente_id=cliente_id, inicio=parse_date_filter(inicio), fim=parse_date_filter(fim), papel=papel, status=status, data_base=base)


@app.get("/relatorios/notas.csv")
def relatorio_notas_csv(
    cliente_id: int | None = None,
    inicio: str | None = None,
    fim: str | None = None,
    papel: str = "PRESTADOR",
    status: str = "TODOS",
    data_base: str = "competencia",
) -> StreamingResponse:
    notes = _notes_for_export(cliente_id, inicio, fim, papel, status, data_base)
    buffer = StringIO()
    buffer.write(";".join(label for _key, label in DETAIL_COLUMNS) + "\n")
    for note in notes:
        row = []
        for key, _label in DETAIL_COLUMNS:
            value = note.get(key, "")
            if key in MONEY_FIELDS:
                row.append(fmt_num(value))
            elif key in {"data_emissao", "competencia"}:
                row.append(fmt_date(str(value or "")))
            elif key in {"cnpj_prestador", "cnpj_tomador"}:
                row.append(mask_cnpj(str(value or "")))
            elif key == "iss_retido":
                row.append("Sim" if value is True else "Não" if value is False else "")
            else:
                row.append(str(value or "").replace(";", ",").replace("\n", " "))
        buffer.write(";".join(row) + "\n")
    return StreamingResponse(
        iter([buffer.getvalue().encode("utf-8-sig")]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=ora_nfse_detalhado.csv"},
    )


@app.get("/relatorios/notas.xlsx")
def relatorio_notas_xlsx(
    cliente_id: int | None = None,
    inicio: str | None = None,
    fim: str | None = None,
    papel: str = "PRESTADOR",
    status: str = "TODOS",
    data_base: str = "competencia",
) -> StreamingResponse:
    data = store.read()
    base = normalize_date_base(data_base)
    notes = filtered_notes(data, cliente_id=cliente_id, inicio=parse_date_filter(inicio), fim=parse_date_filter(fim), papel=papel, status=status, data_base=base)
    totals = totals_for_notes(notes)
    months = monthly_rows(notes, base)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["ORA NFS-e Automático"])
    ws.append(["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws.append(["Cliente", cliente_label(find_by_id(data.get("clientes", []), cliente_id))])
    ws.append(["Período", periodo_label(inicio, fim, base)])
    ws.append(["Data-base", date_base_label(base)])
    ws.append([])
    ws.append(["Totalizador", "Valor"])
    for label, key in [
        ("Notas no filtro", "qtd_total"),
        ("Autorizadas", "qtd_autorizadas"),
        ("Canceladas", "qtd_canceladas"),
        ("Substituídas", "qtd_substituidas"),
        ("Faturamento bruto", "faturamento_bruto"),
        ("Retenções no faturamento", "retencoes_faturamento"),
        ("Valor líquido do faturamento", "faturamento_liquido"),
        ("Valor serviços autorizados", "valor_servico"),
        ("ISS retido", "valor_iss_retido"),
        ("PIS retido", "valor_pis"),
        ("COFINS retido", "valor_cofins"),
        ("INSS/CP retido", "valor_inss"),
        ("IRRF retido", "valor_ir"),
        ("CSLL retida", "valor_csll"),
        ("PIS/COFINS/CSLL agregado no XML", "retencao_pis_cofins_csll_base"),
        ("PIS apurado XML — não retido", "valor_pis_apurado"),
        ("COFINS apurado XML — não retido", "valor_cofins_apurado"),
        ("Outras retenções", "outras_retencoes"),
        ("Total retido", "total_retencoes"),
        ("Valor líquido", "valor_liquido"),
    ]:
        value = totals[key]
        ws.append([label, float(value) if isinstance(value, Decimal) else value])

    ws2 = wb.create_sheet("Resumo mensal")
    ws2.append(["Mês", "Notas autorizadas", "Canceladas", "Serviços", "Retenções", "Valor líquido"])
    for row in months:
        ws2.append([row["mes"], row["qtd_autorizadas"], row["qtd_canceladas"], float(row["faturamento_bruto"]), float(row["retencoes_faturamento"]), float(row["faturamento_liquido"])])

    ws3 = wb.create_sheet("Notas detalhadas")
    ws3.append([label for _key, label in DETAIL_COLUMNS])
    for note in notes:
        row = []
        for key, _label in DETAIL_COLUMNS:
            value = note.get(key)
            if key in MONEY_FIELDS:
                row.append(float(dec(value)))
            elif key in {"data_emissao", "competencia"}:
                row.append(fmt_date(str(value or "")))
            elif key in {"cnpj_prestador", "cnpj_tomador"}:
                row.append(mask_cnpj(str(value or "")))
            elif key == "iss_retido":
                row.append("Sim" if value is True else "Não" if value is False else "")
            elif key == "iss_retido_tipo":
                row.append(iss_retido_tipo_label(value))
            elif key == "retencao_pis_cofins_csll_tipo":
                row.append(social_retencao_tipo_label(value))
            else:
                row.append(value)
        ws3.append(row)

    ws4 = wb.create_sheet("Retenções")
    ret_cols = [
        ("competencia", "Competência"),
        ("numero", "Número"),
        ("razao_prestador", "Prestador"),
        ("razao_tomador", "Tomador"),
        ("papel_cliente", "Papel"),
        ("valor_servico", "Valor serviços"),
        ("valor_iss_retido", "ISS retido"),
        ("iss_retido_tipo", "Tipo retenção ISS"),
        ("valor_pis", "PIS retido"),
        ("valor_cofins", "COFINS retido"),
        ("valor_csll", "CSLL retida"),
        ("valor_ir", "IRRF retido"),
        ("valor_inss", "INSS/CP retido"),
        ("retencao_pis_cofins_csll_tipo", "Tipo PIS/COFINS/CSLL"),
        ("retencao_pis_cofins_csll_base", "PIS/COFINS/CSLL agregado XML"),
        ("retencao_pis_cofins_csll_criterio", "Critério social"),
        ("outras_retencoes", "Outras retenções"),
        ("total_retencoes", "Total retido"),
        ("valor_liquido", "Valor líquido"),
    ]
    ws4.append([label for _key, label in ret_cols])
    for note in notes:
        excel_row = []
        for key, _label in ret_cols:
            if key in MONEY_FIELDS:
                excel_row.append(float(dec(note.get(key))))
            elif key == "competencia":
                excel_row.append(fmt_date(note.get(key)))
            elif key == "iss_retido_tipo":
                excel_row.append(iss_retido_tipo_label(note.get(key)))
            elif key == "retencao_pis_cofins_csll_tipo":
                excel_row.append(social_retencao_tipo_label(note.get(key)))
            else:
                excel_row.append(note.get(key))
        ws4.append(excel_row)

    for sheet in wb.worksheets:
        if sheet.max_row:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="061F3F")
        for col in range(1, min(sheet.max_column, 20) + 1):
            letter = get_column_letter(col)
            sheet.column_dimensions[letter].width = min(max(12, max((len(str(sheet.cell(row=r, column=col).value or "")) for r in range(1, min(sheet.max_row, 80) + 1)), default=12) + 2), 48)
        sheet.freeze_panes = "A2"
        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ora_nfse_relatorio_detalhado.xlsx"},
    )



# ---------------------------------------------------------------------------
# Conferência de planilha externa
# ---------------------------------------------------------------------------


def imported_value_html(value: Any) -> str:
    return escape(str(value or ""))


def render_field_guide_table() -> str:
    rows = "".join(
        "<tr>"
        f"<td><code>{escape(col['cabecalho'])}</code><br><span class='small muted'>{escape(col['campo'])}</span></td>"
        f"<td>{escape(col['obrigatorio'])}</td>"
        f"<td>{escape(col['formato'])}</td>"
        f"<td>{escape(col['orientacao'])}</td>"
        "</tr>"
        for col in EXPECTED_COLUMNS
    )
    return f"<div class='table-wrap guide-table'><table><thead><tr><th>Campo esperado</th><th>Uso</th><th>Formato</th><th>Orientação</th></tr></thead><tbody>{rows}</tbody></table></div>"


def render_comparison_results(result: dict[str, Any] | None) -> str:
    if not result:
        return ""

    avisos = result.get("avisos") or []
    avisos_html = ""
    if avisos:
        avisos_html = "<section class='notice rate-limit-note'><strong>Atenções da importação</strong><span>" + "<br>".join(escape(str(a)) for a in avisos[:12]) + "</span></section>"

    totais = result.get("totais", {})
    kpis = f"""
      <div class="grid five comparison-kpis">
        <div class="card metric blue"><span class="metric-label">Linhas lidas</span><strong class="metric-value">{int(result.get('linhas_lidas') or 0)}</strong><span class="metric-detail">linhas úteis da planilha</span></div>
        <div class="card metric"><span class="metric-label">No sistema</span><strong class="metric-value">{int(result.get('notas_sistema') or 0)}</strong><span class="metric-detail">notas dentro do filtro</span></div>
        <div class="card metric"><span class="metric-label">Conferidas</span><strong class="metric-value">{int(totais.get('conferidas') or 0)}</strong><span class="metric-detail">sem divergência nos campos preenchidos</span></div>
        <div class="card metric accent"><span class="metric-label">Divergências</span><strong class="metric-value">{int(totais.get('divergencias') or 0)}</strong><span class="metric-detail">campos com diferença</span></div>
        <div class="card metric dark"><span class="metric-label">Fora do par</span><strong class="metric-value">{int(totais.get('nao_localizadas') or 0) + int(totais.get('ausentes_planilha') or 0)}</strong><span class="metric-detail">planilha sem sistema ou sistema sem planilha</span></div>
      </div>
    """

    divergencias = result.get("divergencias") or []
    diff_rows = "".join(
        "<tr>"
        f"<td class='num'>{escape(str(item.get('linha') or ''))}</td>"
        f"<td class='small'>{escape(str(item.get('identificador') or ''))}</td>"
        f"<td><strong>{escape(str(item.get('campo') or ''))}</strong><br><span class='small muted'>{escape(str(item.get('motivo') or ''))}</span></td>"
        f"<td>{escape(str(item.get('planilha') or ''))}</td>"
        f"<td>{escape(str(item.get('sistema') or ''))}</td>"
        f"<td>{escape(str(item.get('diferenca') or ''))}</td>"
        "</tr>"
        for item in divergencias[:500]
    ) or "<tr><td colspan='6'>Nenhuma divergência de campo encontrada.</td></tr>"
    if len(divergencias) > 500:
        diff_rows += f"<tr><td colspan='6'>Exibindo as primeiras 500 divergências de {len(divergencias)}.</td></tr>"

    nao_localizadas = result.get("nao_localizadas") or []
    missing_rows = "".join(
        "<tr>"
        f"<td class='num'>{escape(str(item.get('linha') or ''))}</td>"
        f"<td>{escape(str(item.get('identificador') or ''))}</td>"
        f"<td>{escape(str(item.get('motivo') or ''))}</td>"
        "</tr>"
        for item in nao_localizadas[:300]
    ) or "<tr><td colspan='3'>Todas as linhas importadas foram localizadas no sistema.</td></tr>"

    ausentes = result.get("ausentes_planilha") or []
    absent_rows = "".join(
        "<tr>"
        f"<td class='small'>{escape(str(item.get('identificador') or ''))}</td>"
        f"<td>{escape(str(item.get('numero') or ''))}</td>"
        f"<td>{escape(str(item.get('prestador') or ''))}</td>"
        f"<td>{escape(str(item.get('tomador') or ''))}</td>"
        f"<td>{fmt_date(str(item.get('competencia') or ''))}</td>"
        f"<td>{fmt_date(str(item.get('data_emissao') or ''))}</td>"
        f"<td class='num'>{fmt_money(item.get('valor_servico'))}</td>"
        "</tr>"
        for item in ausentes[:300]
    ) or "<tr><td colspan='7'>Nenhuma nota do sistema ficou fora da planilha dentro do filtro.</td></tr>"

    ok_items = result.get("conferidas") or []
    ok_rows = "".join(
        "<tr>"
        f"<td class='num'>{escape(str(item.get('linha') or ''))}</td>"
        f"<td>{escape(str(item.get('identificador') or ''))}</td>"
        f"<td>{escape(str(item.get('metodo') or ''))}</td>"
        "</tr>"
        for item in ok_items[:120]
    ) or "<tr><td colspan='3'>Nenhuma linha ficou totalmente conferida.</td></tr>"

    return f"""
      <section class="comparison-result">
        <div class="section-title"><h2>Resultado da conferência</h2><span class="muted small">campos em branco na planilha não são comparados</span></div>
        {avisos_html}
        {kpis}
        <section class="card result-card">
          <h2>Divergências encontradas</h2>
          <p class="muted small">Diferenças por campo entre a planilha importada e o que foi puxado pelo sistema.</p>
          <div class="table-wrap"><table><thead><tr><th class="num">Linha</th><th>Nota</th><th>Campo</th><th>Planilha</th><th>Sistema</th><th>Diferença</th></tr></thead><tbody>{diff_rows}</tbody></table></div>
        </section>
        <div class="grid two">
          <section class="card result-card">
            <h2>Está na planilha, mas não foi localizado no sistema</h2>
            <div class="table-wrap"><table><thead><tr><th class="num">Linha</th><th>Identificador</th><th>Motivo</th></tr></thead><tbody>{missing_rows}</tbody></table></div>
          </section>
          <section class="card result-card">
            <h2>Está no sistema, mas não veio na planilha</h2>
            <div class="table-wrap"><table><thead><tr><th>Chave/identificador</th><th>Número</th><th>Prestador</th><th>Tomador</th><th>Competência</th><th>Emissão</th><th class="num">Serviços</th></tr></thead><tbody>{absent_rows}</tbody></table></div>
          </section>
        </div>
        <section class="card result-card">
          <h2>Linhas conferidas</h2>
          <p class="muted small">Exibição limitada para manter a leitura leve. Use as divergências acima como trilha principal de correção.</p>
          <div class="table-wrap compact"><table><thead><tr><th class="num">Linha</th><th>Identificador</th><th>Método de localização</th></tr></thead><tbody>{ok_rows}</tbody></table></div>
        </section>
      </section>
    """


def render_conferencia_excel_page(
    data: dict[str, Any],
    result: dict[str, Any] | None = None,
    error: str | None = None,
    cliente_id: int | None = None,
    inicio: str | None = None,
    fim: str | None = None,
    papel: str = "TODOS",
    status: str = "TODOS",
    data_base: str = "competencia",
    tolerancia: str = "0,01",
) -> str:
    base = normalize_date_base(data_base)
    cliente_opts = [("", "Todos os clientes")] + [(c["id"], f"{c.get('razao_social','')} · {mask_cnpj(c.get('cnpj'))}") for c in data.get("clientes", [])]
    papel_opts = [("TODOS", "Emitidas e tomadas"), ("PRESTADOR", "Somente prestadas"), ("TOMADOR", "Somente tomadas")]
    status_opts = [("TODOS", "Todos os status"), ("AUTORIZADA", "Autorizadas"), ("CANCELADA", "Canceladas"), ("SUBSTITUIDA", "Substituídas")]
    data_base_select = date_base_options_html(base)
    error_html = f"<section class='notice danger-notice'><strong>Não foi possível importar a planilha.</strong><span>{escape(error)}</span></section>" if error else ""
    guide = render_field_guide_table()
    result_html = render_comparison_results(result)

    body = f"""
      <section class="ops-layout import-grid">
        <section class="card command-form-card">
          <div class="card-head">
            <div>
              <span class="small-eyebrow">Importação Excel</span>
              <h2>Conferir planilha externa</h2>
            </div>
            <a class="button ghost compact-btn" href="/conferencia-excel/modelo.xlsx">Modelo ORA</a>
          </div>
          <p class="muted small">Compare a planilha do ERP, cliente ou controle interno com as NFS-e importadas no sistema.</p>
          <form method="post" action="/ui/conferencia-excel" enctype="multipart/form-data" class="excel-import-form">
            <div class="form-row two-cols">
              <div><label>Arquivo Excel</label><input type="file" name="arquivo" accept=".xlsx,.xlsm" required></div>
              <div><label>Tolerância monetária</label><input name="tolerancia" value="{escape(tolerancia)}" placeholder="0,01"></div>
            </div>
            <div class="form-row report-filter-row">
              <div><label>Empresa</label><select name="cliente_id">{html_select_options(cliente_opts, cliente_id or '')}</select></div>
              <div><label>Início</label><input type="date" name="inicio" value="{escape(inicio or '')}"></div>
              <div><label>Fim</label><input type="date" name="fim" value="{escape(fim or '')}"></div>
              <div><label>Data-base</label><select name="data_base">{data_base_select}</select></div>
              <div><label>Papel</label><select name="papel">{html_select_options(papel_opts, papel)}</select></div>
              <div><label>Status</label><select name="status">{html_select_options(status_opts, status)}</select></div>
            </div>
            <div class="command-actions">
              <button class="blue" type="submit">Conferir planilha</button>
              <a class="button ghost" href="/conferencia-excel">Limpar</a>
            </div>
          </form>
        </section>

        <aside class="card ops-aside import-rules">
          <div class="card-head">
            <div><span class="small-eyebrow">Regras de importação</span><h2>Como preparar o Excel</h2></div>
          </div>
          <div class="system-checklist">
            <div><strong>1</strong><span>Use a coluna <b>Chave de acesso</b> sempre que possível.</span></div>
            <div><strong>2</strong><span>Sem chave, informe Número + CNPJ prestador + CNPJ tomador.</span></div>
            <div><strong>3</strong><span>Campos em branco não são comparados; campos preenchidos geram conferência.</span></div>
            <div><strong>4</strong><span>Valores aceitam 1234,56 ou R$ 1.234,56; datas aceitam Excel ou dd/mm/aaaa.</span></div>
          </div>
        </aside>
      </section>

      {error_html}
      {result_html}

      <section class="notice system-note">
        <strong>Orientação de uso</strong>
        <span>O filtro acima define quais notas do sistema entram na comparação. Notas fora do filtro podem aparecer como não localizadas.</span>
      </section>

      <section class="card data-card guide-card">
        <div class="card-head">
          <div><span class="small-eyebrow">Cabeçalhos aceitos</span><h2>Campos da planilha</h2></div>
          <span class="muted small">A primeira aba deve conter o cabeçalho até a linha 12.</span>
        </div>
        {guide}
      </section>
    """
    return render_page("Conferência Excel", "conferencia", body, subtitle="Importação e comparação de planilha externa contra a base local.")

@app.get("/conferencia-excel", response_class=HTMLResponse)
def conferencia_excel(
    cliente_id: int | None = Query(default=None),
    inicio: str | None = Query(default=None),
    fim: str | None = Query(default=None),
    papel: str = Query(default="TODOS"),
    status: str = Query(default="TODOS"),
    data_base: str = Query(default="competencia"),
    tolerancia: str = Query(default="0,01"),
) -> str:
    data = store.read()
    return render_conferencia_excel_page(
        data,
        cliente_id=cliente_id,
        inicio=inicio,
        fim=fim,
        papel=papel,
        status=status,
        data_base=data_base,
        tolerancia=tolerancia,
    )


@app.get("/conferencia-excel/modelo.xlsx")
def modelo_conferencia_excel() -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Importação"
    headers = [col["cabecalho"] for col in EXPECTED_COLUMNS]
    ws.append(headers)
    ws.freeze_panes = "A2"

    example = wb.create_sheet("Exemplo")
    example.append(headers)
    example.append([
        "12345678901234567890123456789012345678901234",
        "1001",
        "11.111.111/0001-91",
        "22.222.222/0001-82",
        "01/02/2026",
        "05/02/2026",
        "PRESTADOR",
        "AUTORIZADA",
        10000.00,
        10000.00,
        500.00,
        500.00,
        65.00,
        300.00,
        100.00,
        150.00,
        0.00,
        0.00,
        1115.00,
        8885.00,
    ])

    guide = wb.create_sheet("Orientações")
    guide.append(["Campo", "Uso", "Formato", "Orientação"])
    for col in EXPECTED_COLUMNS:
        guide.append([col["cabecalho"], col["obrigatorio"], col["formato"], col["orientacao"]])
    guide.append([])
    guide.append(["Regra", "Campo em branco não é comparado. Para localizar a nota, use Chave de acesso ou Número + CNPJ prestador + CNPJ tomador."])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="061F3F")
        for col in range(1, min(sheet.max_column, 28) + 1):
            letter = get_column_letter(col)
            sheet.column_dimensions[letter].width = min(max(14, max((len(str(sheet.cell(row=r, column=col).value or "")) for r in range(1, min(sheet.max_row, 50) + 1)), default=14) + 2), 42)
        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ora_modelo_conferencia_nfse.xlsx"},
    )


@app.post("/ui/conferencia-excel", response_class=HTMLResponse)
async def ui_conferencia_excel(
    arquivo: UploadFile = File(...),
    cliente_id: int | None = Form(default=None),
    inicio: str | None = Form(default=None),
    fim: str | None = Form(default=None),
    papel: str = Form(default="TODOS"),
    status: str = Form(default="TODOS"),
    data_base: str = Form(default="competencia"),
    tolerancia: str = Form(default="0,01"),
) -> str:
    data = store.read()
    filename = arquivo.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        return render_conferencia_excel_page(data, error="Envie uma planilha Excel no formato .xlsx ou .xlsm.", cliente_id=cliente_id, inicio=inicio, fim=fim, papel=papel, status=status, data_base=data_base, tolerancia=tolerancia)

    content = await arquivo.read()
    if len(content) > 12 * 1024 * 1024:
        return render_conferencia_excel_page(data, error="O arquivo excede 12 MB. Divida a conferência em planilhas menores.", cliente_id=cliente_id, inicio=inicio, fim=fim, papel=papel, status=status, data_base=data_base, tolerancia=tolerancia)

    try:
        imported_rows, import_warnings, _headers = read_excel_rows(content)
    except ValueError as exc:
        return render_conferencia_excel_page(data, error=str(exc), cliente_id=cliente_id, inicio=inicio, fim=fim, papel=papel, status=status, data_base=data_base, tolerancia=tolerancia)

    base = normalize_date_base(data_base)
    notes = filtered_notes(
        data,
        cliente_id=cliente_id,
        inicio=parse_date_filter(inicio),
        fim=parse_date_filter(fim),
        papel=papel,
        status=status,
        data_base=base,
    )
    tolerance_value = parse_excel_money(tolerancia) or Decimal("0.01")
    result = compare_excel_with_notes(imported_rows, notes, tolerance=tolerance_value)
    result["avisos"] = import_warnings + list(result.get("avisos") or [])

    return render_conferencia_excel_page(
        data,
        result=result,
        cliente_id=cliente_id,
        inicio=inicio,
        fim=fim,
        papel=papel,
        status=status,
        data_base=base,
        tolerancia=tolerancia,
    )


@app.get("/sincronizacao/logs", response_class=HTMLResponse)
def logs() -> str:
    data = store.read()
    logs = sorted(data.get("logs", []), key=lambda x: x.get("criado_em", ""), reverse=True)[:100]
    rows = []
    for log in logs:
        msgs = [str(m) for m in (log.get("mensagens", []) or [])[-4:]]
        msgs_html = "<br>".join(escape(m) for m in msgs) or "-"
        status = str(log.get("status") or "")
        origem = str(log.get("origem") or "ADN_API")
        cnpj = mask_cnpj(log.get("cnpj_cliente")) if log.get("cnpj_cliente") else "-"
        rows.append(
            "<tr>"
            f"<td>{log.get('id')}</td>"
            f"<td>{escape(origem)}</td>"
            f"<td>{cnpj}</td>"
            f"<td>{pill(status, 'ok' if status == 'CONCLUIDA' else 'err' if status == 'ERRO' else 'warn')}</td>"
            f"<td>{escape(str(log.get('criado_em','')).replace('T',' ')[:19])}</td>"
            f"<td class='num'>{log.get('consultas_realizadas', '')}</td>"
            f"<td class='num'>{log.get('documentos_recebidos', '')}</td>"
            f"<td class='num'>{log.get('importadas', '')}</td>"
            f"<td class='num'>{log.get('atualizadas', '')}</td>"
            f"<td>{msgs_html}</td>"
            "</tr>"
        )
    body_rows = "".join(rows) or "<tr><td colspan='10'>Nenhuma busca registrada ainda.</td></tr>"
    html = f"""
      <section class="toolbar-card">
        <div class="toolbar-copy">
          <span class="small-eyebrow">Rastreabilidade</span>
          <h2>Histórico técnico</h2>
          <p>Registro das últimas buscas, importações e reprocessamentos para auditoria e suporte.</p>
        </div>
        <div class="toolbar-stats">
          <div><span>Registros</span><strong>{len(logs)}</strong></div>
          <div><span>Concluídos</span><strong>{sum(1 for item in logs if item.get('status') == 'CONCLUIDA')}</strong></div>
          <div><span>Alertas/erros</span><strong>{sum(1 for item in logs if item.get('status') != 'CONCLUIDA')}</strong></div>
        </div>
      </section>
      <section class="card data-card">
        <div class="card-head"><div><span class="small-eyebrow">Operações</span><h2>Últimos registros</h2></div></div>
        <div class="table-wrap"><table><thead><tr><th>ID</th><th>Origem</th><th>CNPJ</th><th>Status</th><th>Criado em</th><th class="num">Consultas</th><th class="num">Docs</th><th class="num">Importadas</th><th class="num">Atualizadas</th><th>Mensagens úteis</th></tr></thead><tbody>{body_rows}</tbody></table></div>
      </section>
    """
    return render_page("Histórico", "logs", html, subtitle="Logs técnicos sem poluir a rotina operacional.")

# ---------------------------------------------------------------------------
# API simples
# ---------------------------------------------------------------------------


@app.get("/api/clientes")
def api_clientes() -> list[dict[str, Any]]:
    data = store.read()
    return data.get("clientes", [])


@app.get("/api/notas")
def api_notas(
    cliente_id: int | None = None,
    inicio: str | None = None,
    fim: str | None = None,
    papel: str = "PRESTADOR",
    status: str = "TODOS",
    data_base: str = "competencia",
) -> list[dict[str, Any]]:
    data = store.read()
    return filtered_notes(data, cliente_id=cliente_id, inicio=parse_date_filter(inicio), fim=parse_date_filter(fim), papel=papel, status=status, data_base=normalize_date_base(data_base))


@app.get("/api/resumo")
def api_resumo(
    cliente_id: int | None = None,
    inicio: str | None = None,
    fim: str | None = None,
    papel: str = "PRESTADOR",
    status: str = "TODOS",
    data_base: str = "competencia",
) -> dict[str, Any]:
    data = store.read()
    notes = filtered_notes(data, cliente_id=cliente_id, inicio=parse_date_filter(inicio), fim=parse_date_filter(fim), papel=papel, status=status, data_base=normalize_date_base(data_base))
    totals = totals_for_notes(notes)
    return {key: (str(value) if isinstance(value, Decimal) else value) for key, value in totals.items()}


def json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, list):
        return [json_safe(v) for v in value]
    if isinstance(value, dict):
        return {k: json_safe(v) for k, v in value.items()}
    return value


@app.get("/api/resumo-empresas")
def api_resumo_empresas(
    inicio: str | None = None,
    fim: str | None = None,
    status: str = "TODOS",
    data_base: str = "competencia",
    cliente_ids: list[int] = Query(default=[]),
) -> dict[str, Any]:
    data = store.read()
    rows = resumo_empresas_rows(data, inicio=parse_date_filter(inicio), fim=parse_date_filter(fim), status=status, cliente_ids=normalize_id_list(cliente_ids), data_base=normalize_date_base(data_base))
    totals = resumo_empresas_totais(rows)
    return json_safe({"totais": totals, "empresas": rows})


@app.delete("/api/certificados/{certificado_id}")
def api_excluir_certificado(certificado_id: int) -> dict[str, Any]:
    return excluir_certificado(certificado_id)


@app.delete("/api/clientes/{cliente_id}")
def api_excluir_cliente(cliente_id: int) -> dict[str, Any]:
    return excluir_cliente(cliente_id)
