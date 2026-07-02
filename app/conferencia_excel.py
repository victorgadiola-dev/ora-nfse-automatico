from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

ZERO = Decimal("0.00")

EXPECTED_COLUMNS: list[dict[str, str]] = [
    {"campo": "chave_acesso", "cabecalho": "Chave de acesso", "obrigatorio": "recomendado", "formato": "Texto/número", "orientacao": "Melhor campo para localizar a nota sem ambiguidade."},
    {"campo": "numero", "cabecalho": "Número", "obrigatorio": "condicional", "formato": "Texto/número", "orientacao": "Obrigatório quando a planilha não tiver chave de acesso."},
    {"campo": "cnpj_prestador", "cabecalho": "CNPJ prestador", "obrigatorio": "condicional", "formato": "00.000.000/0000-00", "orientacao": "Obrigatório quando a busca for feita por número."},
    {"campo": "cnpj_tomador", "cabecalho": "CNPJ tomador", "obrigatorio": "condicional", "formato": "00.000.000/0000-00", "orientacao": "Obrigatório quando a busca for feita por número."},
    {"campo": "competencia", "cabecalho": "Competência", "obrigatorio": "opcional", "formato": "dd/mm/aaaa", "orientacao": "Será comparada apenas se preenchida."},
    {"campo": "data_emissao", "cabecalho": "Data de emissão", "obrigatorio": "opcional", "formato": "dd/mm/aaaa", "orientacao": "Será comparada apenas se preenchida."},
    {"campo": "papel_cliente", "cabecalho": "Papel", "obrigatorio": "opcional", "formato": "PRESTADOR ou TOMADOR", "orientacao": "Ajuda a identificar se é nota prestada ou tomada."},
    {"campo": "status", "cabecalho": "Status", "obrigatorio": "opcional", "formato": "AUTORIZADA/CANCELADA/SUBSTITUIDA", "orientacao": "Divergências de status aparecem como alerta relevante."},
    {"campo": "valor_servico", "cabecalho": "Valor dos serviços", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Use número ou moeda brasileira. Ex.: 1234,56."},
    {"campo": "base_calculo", "cabecalho": "Base ISS", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Base de cálculo do ISS."},
    {"campo": "valor_iss", "cabecalho": "ISS destacado", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Valor do ISS informado/destacado no XML."},
    {"campo": "valor_iss_retido", "cabecalho": "ISS retido", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Retenção de ISS."},
    {"campo": "valor_pis", "cabecalho": "PIS retido", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Retenção de PIS. Não confundir com PIS apurado no XML."},
    {"campo": "valor_cofins", "cabecalho": "COFINS retido", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Retenção de COFINS. Não confundir com COFINS apurado no XML."},
    {"campo": "valor_csll", "cabecalho": "CSLL retida", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Retenção de CSLL."},
    {"campo": "valor_ir", "cabecalho": "IRRF retido", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Retenção de IRRF."},
    {"campo": "valor_inss", "cabecalho": "INSS/CP retido", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Retenção previdenciária/CP."},
    {"campo": "outras_retencoes", "cabecalho": "Outras retenções", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Outras retenções que a planilha controle separadamente."},
    {"campo": "total_retencoes", "cabecalho": "Total retido", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Soma das retenções. Se preenchido, o sistema compara com o total calculado."},
    {"campo": "valor_liquido", "cabecalho": "Valor líquido", "obrigatorio": "opcional", "formato": "Moeda", "orientacao": "Valor líquido da NFS-e."},
]

MONEY_FIELDS = {
    "valor_servico",
    "base_calculo",
    "valor_iss",
    "valor_iss_retido",
    "valor_pis",
    "valor_cofins",
    "valor_csll",
    "valor_ir",
    "valor_inss",
    "outras_retencoes",
    "total_retencoes",
    "valor_liquido",
}
DATE_FIELDS = {"competencia", "data_emissao"}
TEXT_FIELDS = {"numero", "papel_cliente", "status", "cnpj_prestador", "cnpj_tomador"}

COMPARE_FIELDS: list[tuple[str, str, str]] = [
    ("numero", "Número", "text"),
    ("cnpj_prestador", "CNPJ prestador", "cnpj"),
    ("cnpj_tomador", "CNPJ tomador", "cnpj"),
    ("competencia", "Competência", "date"),
    ("data_emissao", "Data de emissão", "date"),
    ("papel_cliente", "Papel", "text"),
    ("status", "Status", "text"),
    ("valor_servico", "Valor dos serviços", "money"),
    ("base_calculo", "Base ISS", "money"),
    ("valor_iss", "ISS destacado", "money"),
    ("valor_iss_retido", "ISS retido", "money"),
    ("valor_pis", "PIS retido", "money"),
    ("valor_cofins", "COFINS retido", "money"),
    ("valor_csll", "CSLL retida", "money"),
    ("valor_ir", "IRRF retido", "money"),
    ("valor_inss", "INSS/CP retido", "money"),
    ("outras_retencoes", "Outras retenções", "money"),
    ("total_retencoes", "Total retido", "money"),
    ("valor_liquido", "Valor líquido", "money"),
]

ALIASES = {
    "chave": "chave_acesso",
    "chaveacesso": "chave_acesso",
    "chavedeacesso": "chave_acesso",
    "chavenfse": "chave_acesso",
    "chavedanfse": "chave_acesso",
    "chavenfsenacional": "chave_acesso",
    "codigoacesso": "chave_acesso",
    "numero": "numero",
    "n": "numero",
    "nfse": "numero",
    "numeronfse": "numero",
    "nnfse": "numero",
    "cnpjprestador": "cnpj_prestador",
    "prestadorcnpj": "cnpj_prestador",
    "documentoprestador": "cnpj_prestador",
    "cnpjdo prestador": "cnpj_prestador",
    "cnpjtomador": "cnpj_tomador",
    "tomadorcnpj": "cnpj_tomador",
    "documentotomador": "cnpj_tomador",
    "cnpjdo tomador": "cnpj_tomador",
    "competencia": "competencia",
    "dtcompetencia": "competencia",
    "datacompetencia": "competencia",
    "data competencia": "competencia",
    "emissao": "data_emissao",
    "dataemissao": "data_emissao",
    "datadeemissao": "data_emissao",
    "dtemissao": "data_emissao",
    "papel": "papel_cliente",
    "natureza": "papel_cliente",
    "papelcliente": "papel_cliente",
    "tipo": "papel_cliente",
    "status": "status",
    "situacao": "status",
    "valorservico": "valor_servico",
    "valordosservicos": "valor_servico",
    "valorservicos": "valor_servico",
    "vserv": "valor_servico",
    "baseiss": "base_calculo",
    "basecalculo": "base_calculo",
    "basedecalculo": "base_calculo",
    "bciss": "base_calculo",
    "iss": "valor_iss",
    "issdestacado": "valor_iss",
    "valoriss": "valor_iss",
    "viss": "valor_iss",
    "issretido": "valor_iss_retido",
    "valorissretido": "valor_iss_retido",
    "vissret": "valor_iss_retido",
    "pisretido": "valor_pis",
    "valorpisretido": "valor_pis",
    "retencaopis": "valor_pis",
    "cofinsretido": "valor_cofins",
    "valorcofinsretido": "valor_cofins",
    "retencaocofins": "valor_cofins",
    "csllretida": "valor_csll",
    "csllretido": "valor_csll",
    "valorcsllretida": "valor_csll",
    "retencaocsll": "valor_csll",
    "irrfretido": "valor_ir",
    "irretido": "valor_ir",
    "valorirrfretido": "valor_ir",
    "valorirretido": "valor_ir",
    "retencaoir": "valor_ir",
    "inssretido": "valor_inss",
    "insscpretido": "valor_inss",
    "valorinssretido": "valor_inss",
    "retencaoinss": "valor_inss",
    "outrasretencoes": "outras_retencoes",
    "outrasret": "outras_retencoes",
    "totalretido": "total_retencoes",
    "totalretencoes": "total_retencoes",
    "retencoestotais": "total_retencoes",
    "valorliquido": "valor_liquido",
    "liquido": "valor_liquido",
    "vliquido": "valor_liquido",
}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    replacements = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "õ": "o", "ô": "o",
        "ú": "u",
        "ç": "c",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[^a-z0-9]", "", text)


def canonical_header(value: Any) -> str | None:
    normalized = normalize_header(value)
    if not normalized:
        return None
    if normalized in ALIASES:
        return ALIASES[normalized]
    for col in EXPECTED_COLUMNS:
        if normalized == normalize_header(col["cabecalho"]):
            return col["campo"]
    return None


def clean_digits(value: Any) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def normalize_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def normalize_number(value: Any) -> str:
    cleaned = normalize_key(value)
    if cleaned.isdigit():
        return cleaned.lstrip("0") or "0"
    return cleaned


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def parse_money(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, int | float):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\u00a0", " ")
    cleaned = re.sub(r"[^0-9,.\-]", "", text)
    if not cleaned:
        return None
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def parse_date_value(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("Z", "+00:00")
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value)


def note_identity(note: dict[str, Any]) -> str:
    chave = normalize_key(note.get("chave_acesso"))
    if chave:
        return f"chave:{chave}"
    return "cmp:" + "|".join([
        normalize_number(note.get("numero")),
        clean_digits(note.get("cnpj_prestador")),
        clean_digits(note.get("cnpj_tomador")),
    ])


def row_identity(row: dict[str, Any]) -> str:
    chave = normalize_key(row.get("chave_acesso"))
    if chave:
        return f"chave:{chave}"
    return "cmp:" + "|".join([
        normalize_number(row.get("numero")),
        clean_digits(row.get("cnpj_prestador")),
        clean_digits(row.get("cnpj_tomador")),
    ])


def composite_key(data: dict[str, Any]) -> str:
    return "|".join([
        normalize_key(data.get("numero")),
        clean_digits(data.get("cnpj_prestador")),
        clean_digits(data.get("cnpj_tomador")),
    ])


def import_identifier(row: dict[str, Any]) -> str:
    chave = normalize_key(row.get("chave_acesso"))
    if chave:
        return chave
    numero = row.get("numero") or "sem número"
    prest = clean_digits(row.get("cnpj_prestador")) or "sem prestador"
    toma = clean_digits(row.get("cnpj_tomador")) or "sem tomador"
    return f"NF {numero} · {prest} · {toma}"


def note_identifier(note: dict[str, Any]) -> str:
    chave = normalize_key(note.get("chave_acesso"))
    if chave:
        return chave
    numero = note.get("numero") or "sem número"
    prest = clean_digits(note.get("cnpj_prestador")) or "sem prestador"
    toma = clean_digits(note.get("cnpj_tomador")) or "sem tomador"
    return f"NF {numero} · {prest} · {toma}"


def read_excel_rows(content: bytes, max_rows: int = 20000) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    """Lê a primeira aba do Excel e retorna linhas normalizadas, avisos e cabeçalhos reconhecidos."""
    if not content:
        return [], ["Arquivo vazio."], []
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise ValueError("Não foi possível abrir o Excel. Envie arquivo .xlsx ou .xlsm válido.") from exc

    ws = wb.active
    header_row_idx = None
    headers: list[str | None] = []
    raw_headers: list[str] = []
    warnings: list[str] = []
    recognized_headers: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        canonical = [canonical_header(cell) for cell in row]
        recognized = [c for c in canonical if c]
        if len(recognized) >= 3 and ("chave_acesso" in recognized or "numero" in recognized):
            header_row_idx = row_idx
            headers = canonical
            raw_headers = [str(cell or "").strip() for cell in row]
            recognized_headers = recognized
            break

    if header_row_idx is None:
        raise ValueError("Não encontrei a linha de cabeçalho. Use o modelo ORA ou mantenha cabeçalhos como Chave de acesso, Número, CNPJ prestador, CNPJ tomador, Competência e valores.")

    unknown = [h for h, canon in zip(raw_headers, headers) if h and not canon]
    if unknown:
        warnings.append("Colunas ignoradas: " + ", ".join(unknown[:12]) + ("..." if len(unknown) > 12 else ""))

    rows: list[dict[str, Any]] = []
    for offset, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if offset - header_row_idx > max_rows:
            warnings.append(f"Importação limitada às primeiras {max_rows} linhas úteis.")
            break
        if all(is_blank(v) for v in row):
            continue
        item: dict[str, Any] = {"_linha_excel": offset}
        has_any = False
        for idx, raw in enumerate(row):
            canonical = headers[idx] if idx < len(headers) else None
            if not canonical:
                continue
            if not is_blank(raw):
                has_any = True
            item[canonical] = raw
        if has_any:
            rows.append(item)

    if not rows:
        warnings.append("A planilha não possui linhas úteis após o cabeçalho.")
    if "chave_acesso" not in recognized_headers:
        warnings.append("Sem coluna de chave de acesso: a conciliação usará Número + CNPJ prestador + CNPJ tomador, que pode ser ambígua.")
    return rows, warnings, recognized_headers


def compare_excel_with_notes(rows: list[dict[str, Any]], notes: list[dict[str, Any]], tolerance: Decimal | str | float = Decimal("0.01")) -> dict[str, Any]:
    tol = parse_money(tolerance)
    if tol is None:
        tol = Decimal("0.01")

    chave_index: dict[str, dict[str, Any]] = {}
    composite_index: dict[str, list[dict[str, Any]]] = {}
    for note in notes:
        chave = normalize_key(note.get("chave_acesso"))
        if chave:
            chave_index[chave] = note
        cmp_key = composite_key(note)
        if cmp_key.strip("|"):
            composite_index.setdefault(cmp_key, []).append(note)

    matched_identities: set[int] = set()
    conferidas: list[dict[str, Any]] = []
    divergencias: list[dict[str, Any]] = []
    nao_localizadas: list[dict[str, Any]] = []
    avisos: list[str] = []

    for row in rows:
        linha = row.get("_linha_excel")
        chave = normalize_key(row.get("chave_acesso"))
        match: dict[str, Any] | None = None
        metodo = ""

        if chave:
            match = chave_index.get(chave)
            metodo = "chave de acesso"

        if match is None:
            cmp_key = composite_key(row)
            candidates = composite_index.get(cmp_key, [])
            if len(candidates) == 1:
                match = candidates[0]
                metodo = "número + CNPJs"
            elif len(candidates) > 1:
                avisos.append(f"Linha {linha}: mais de uma nota encontrada para Número + CNPJs. Preencha a chave de acesso para eliminar ambiguidade.")
                nao_localizadas.append({
                    "linha": linha,
                    "identificador": import_identifier(row),
                    "motivo": "Correspondência ambígua no sistema.",
                })
                continue

        if match is None:
            if not chave and not composite_key(row).strip("|"):
                motivo = "Sem chave de acesso ou conjunto Número + CNPJs."
            else:
                motivo = "Não existe nota correspondente dentro do filtro aplicado."
            nao_localizadas.append({
                "linha": linha,
                "identificador": import_identifier(row),
                "motivo": motivo,
            })
            continue

        matched_identities.add(id(match))
        row_issues: list[dict[str, Any]] = []
        for field, label, kind in COMPARE_FIELDS:
            if field not in row or is_blank(row.get(field)):
                continue
            imported_raw = row.get(field)
            system_raw = match.get(field)

            if kind == "money":
                imported = parse_money(imported_raw)
                system = parse_money(system_raw) or ZERO
                if imported is None:
                    row_issues.append({
                        "linha": linha,
                        "identificador": import_identifier(row),
                        "campo": label,
                        "planilha": display_value(imported_raw),
                        "sistema": display_value(system),
                        "diferenca": "",
                        "motivo": "Valor monetário inválido na planilha.",
                    })
                    continue
                diff = imported - system
                if abs(diff) > tol:
                    row_issues.append({
                        "linha": linha,
                        "identificador": import_identifier(row),
                        "campo": label,
                        "planilha": display_value(imported),
                        "sistema": display_value(system),
                        "diferenca": display_value(diff),
                        "motivo": f"Diferença superior à tolerância de {display_value(tol)}.",
                    })
            elif kind == "date":
                imported_date = parse_date_value(imported_raw)
                system_date = parse_date_value(system_raw)
                if imported_date != system_date:
                    row_issues.append({
                        "linha": linha,
                        "identificador": import_identifier(row),
                        "campo": label,
                        "planilha": display_value(imported_date or imported_raw),
                        "sistema": display_value(system_date or system_raw),
                        "diferenca": "",
                        "motivo": "Data diferente.",
                    })
            elif kind == "cnpj":
                imported_doc = clean_digits(imported_raw)
                system_doc = clean_digits(system_raw)
                if imported_doc != system_doc:
                    row_issues.append({
                        "linha": linha,
                        "identificador": import_identifier(row),
                        "campo": label,
                        "planilha": imported_doc,
                        "sistema": system_doc,
                        "diferenca": "",
                        "motivo": "Documento diferente.",
                    })
            else:
                if field == "numero":
                    imported_text = normalize_number(imported_raw)
                    system_text = normalize_number(system_raw)
                else:
                    imported_text = normalize_key(imported_raw)
                    system_text = normalize_key(system_raw)
                if imported_text != system_text:
                    row_issues.append({
                        "linha": linha,
                        "identificador": import_identifier(row),
                        "campo": label,
                        "planilha": display_value(imported_raw),
                        "sistema": display_value(system_raw),
                        "diferenca": "",
                        "motivo": "Informação textual diferente.",
                    })

        if row_issues:
            divergencias.extend(row_issues)
        else:
            conferidas.append({
                "linha": linha,
                "identificador": import_identifier(row),
                "metodo": metodo,
            })

    ausentes_planilha = []
    for note in notes:
        if id(note) in matched_identities:
            continue
        ausentes_planilha.append({
            "identificador": note_identifier(note),
            "numero": note.get("numero") or "",
            "prestador": note.get("razao_prestador") or "",
            "tomador": note.get("razao_tomador") or "",
            "competencia": note.get("competencia") or "",
            "data_emissao": note.get("data_emissao") or "",
            "valor_servico": note.get("valor_servico") or "0.00",
        })

    return {
        "linhas_lidas": len(rows),
        "notas_sistema": len(notes),
        "conferidas": conferidas,
        "divergencias": divergencias,
        "nao_localizadas": nao_localizadas,
        "ausentes_planilha": ausentes_planilha,
        "avisos": avisos,
        "totais": {
            "conferidas": len(conferidas),
            "divergencias": len(divergencias),
            "nao_localizadas": len(nao_localizadas),
            "ausentes_planilha": len(ausentes_planilha),
        },
    }
